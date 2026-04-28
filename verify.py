from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path

from libs import (
    collect_python_packages,
    collect_static_library_projects,
    collect_verification_imports,
    collect_verification_steps,
    load_integrations,
)


def log(message: str) -> None:
    text = f"[staticpython-verify] {message}"
    encoding = sys.stdout.encoding or "utf-8"
    safe_text = text.encode(encoding, errors="backslashreplace").decode(encoding, errors="replace")
    print(safe_text, flush=True)


def load_manifest(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_config(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def resolve_profile(config: dict, profile_name: str | None) -> tuple[str, dict]:
    profiles = config.get("profiles", {})
    selected_name = profile_name or config.get("default_profile") or "full"
    if selected_name not in profiles:
        available = ", ".join(sorted(profiles)) or "<none>"
        raise RuntimeError(f"unknown profile {selected_name!r}; available profiles: {available}")
    profile = profiles[selected_name]
    if not isinstance(profile, dict):
        raise RuntimeError(f"profile {selected_name!r} must be an object")
    return selected_name, profile


def project_exists(source_root: Path, project_name: str) -> bool:
    return (source_root / "PCbuild" / project_name).exists()


def all_static_library_projects(manifest: dict, integrations: list) -> list[str]:
    manifest_projects = manifest.get(
        "static_library_projects_release_x64",
        [project["project"] for project in manifest.get("native_static_projects", [])],
    )
    return list(dict.fromkeys([*manifest_projects, *collect_static_library_projects(integrations)]))


def available_manifest_modules(source_root: Path, manifest: dict, integrations: list) -> list[str]:
    modules = list(
        dict.fromkeys(
            [
                *collect_python_packages(integrations),
                *collect_verification_imports(integrations),
            ]
        )
    )
    available_projects = {
        Path(project).stem
        for project in all_static_library_projects(manifest, integrations)
        if project_exists(source_root, project)
    }
    for module_name in manifest.get("python_builtin_modules", []):
        if module_name in available_projects:
            modules.append(module_name)
        else:
            log(
                f"skip manifest verification import {module_name} because {module_name}.vcxproj "
                "does not exist in this CPython version"
            )
    return list(dict.fromkeys(modules))


def _to_text(value: str | bytes | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return value


def run_capture(cmd: list[str], *, cwd: Path, timeout: float) -> dict:
    display = subprocess.list2cmdline([str(part) for part in cmd])
    log(f"RUN {display}")
    try:
        completed = subprocess.run(
            cmd,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        return {
            "ok": False,
            "timeout": True,
            "returncode": None,
            "stdout": _to_text(exc.stdout),
            "stderr": _to_text(exc.stderr),
            "error": f"timed out after {timeout} seconds",
            "display": display,
        }
    return {
        "ok": completed.returncode == 0,
        "timeout": False,
        "returncode": completed.returncode,
        "stdout": completed.stdout,
        "stderr": completed.stderr,
        "display": display,
    }


def build_import_check_code(source_root: Path, manifest: dict, integrations: list) -> str:
    modules_json = json.dumps(
        available_manifest_modules(source_root, manifest, integrations),
        ensure_ascii=False,
    )
    return f"""
import importlib
import json
import traceback

mods = json.loads({modules_json!r})
results = []

for name in mods:
    try:
        module = importlib.import_module(name)
    except Exception as exc:
        results.append({{
            "name": name,
            "ok": False,
            "error_type": type(exc).__name__,
            "error": str(exc),
            "traceback": traceback.format_exc(),
        }})
    else:
        results.append({{
            "name": name,
            "ok": True,
            "origin": getattr(getattr(module, "__spec__", None), "origin", None),
        }})

print(json.dumps(results, ensure_ascii=False, sort_keys=True))
"""


STDLIB_SMOKE = r"""
import importlib.resources
import io
import json
import struct
import traceback

results = []


def must(condition, message):
    if not condition:
        raise AssertionError(message)


def check(name):
    def decorator(func):
        try:
            func()
        except Exception as exc:
            results.append(
                {
                    "name": name,
                    "ok": False,
                    "error_type": type(exc).__name__,
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                }
            )
        else:
            results.append({"name": name, "ok": True})
        return func

    return decorator


@check("stdlib.core")
def _():
    import bz2
    import ctypes
    import lzma
    import queue
    import select
    import sqlite3
    import socket
    import ssl
    import unicodedata
    import uuid
    import xml.etree.ElementTree as ET
    import zoneinfo
    from copy import deepcopy
    from datetime import datetime, timedelta, timezone
    from decimal import Decimal
    from pathlib import Path

    conn = sqlite3.connect(":memory:")
    conn.execute("create table demo (id integer primary key, name text)")
    conn.execute("insert into demo(name) values (?)", ("sandbox",))
    row = conn.execute("select name from demo").fetchone()
    conn.close()
    must(row == ("sandbox",), "sqlite3 roundtrip failed")
    must(bz2.decompress(bz2.compress(b"demo")) == b"demo", "bz2 roundtrip failed")
    must(lzma.decompress(lzma.compress(b"demo")) == b"demo", "lzma roundtrip failed")
    # The stdlib-only profile intentionally does not include the third-party
    # tzdata package, and Windows runners do not provide an IANA tzdb. Use a
    # tiny embedded UTC TZif file so this verifies zoneinfo itself.
    utc_tzif = (
        b"TZif\0"
        + (b"\0" * 15)
        + struct.pack(">6l", 0, 0, 0, 0, 1, 4)
        + struct.pack(">lbb", 0, 0, 0)
        + b"UTC\0"
    )
    utc_zone = zoneinfo.ZoneInfo.from_file(io.BytesIO(utc_tzif), key="UTC")
    must(utc_zone.key == "UTC", "zoneinfo key failed")
    must(datetime(2024, 1, 1, tzinfo=utc_zone).utcoffset() == timedelta(0), "zoneinfo failed")
    must(unicodedata.lookup("LATIN CAPITAL LETTER A") == "A", "unicodedata failed")
    sock_a, sock_b = socket.socketpair()
    try:
        ready, _, _ = select.select([sock_a], [], [], 0)
        must(ready == [], "select failed")
    finally:
        sock_a.close()
        sock_b.close()
    must(ctypes.sizeof(ctypes.c_void_p) in (4, 8), "ctypes failed")
    root = ET.fromstring("<root><child value='1'/></root>")
    must(root[0].attrib["value"] == "1", "elementtree failed")
    q = queue.SimpleQueue()
    q.put("ok")
    must(q.get() == "ok", "queue.SimpleQueue failed")
    ctx = ssl.create_default_context()
    must(hasattr(ctx, "check_hostname"), "ssl context failed")
    must(uuid.UUID(str(uuid.uuid4())).version == 4, "uuid failed")
    must(Decimal("1.1") + Decimal("2.2") == Decimal("3.3"), "decimal failed")
    must(deepcopy({"items": [1, 2, 3]}) == {"items": [1, 2, 3]}, "deepcopy failed")
    must(
        datetime.now(timezone.utc) + timedelta(seconds=1) > datetime.now(timezone.utc),
        "datetime failed",
    )
    must(str(Path("demo") / "child") == "demo\\child", "pathlib failed")


print(json.dumps(results, ensure_ascii=False, sort_keys=True))
"""


THIRD_PARTY_SMOKE = r"""
import asyncio
import importlib.resources
import io
import json
import traceback

results = []


def must(condition, message):
    if not condition:
        raise AssertionError(message)


def check(name):
    def decorator(func):
        try:
            func()
        except Exception as exc:
            results.append(
                {
                    "name": name,
                    "ok": False,
                    "error_type": type(exc).__name__,
                    "error": str(exc),
                    "traceback": traceback.format_exc(),
                }
            )
        else:
            results.append({"name": name, "ok": True})
        return func

    return decorator


@check("pycryptodome.aes")
def _():
    from Crypto.Cipher import AES
    from Crypto.Random import get_random_bytes
    from Crypto.Util.Padding import pad, unpad

    data = b"singlefile-runtime"
    key = b"0123456789abcdef"
    iv = b"abcdef9876543210"
    ciphertext = AES.new(key, AES.MODE_CBC, iv).encrypt(pad(data, 16))
    plaintext = unpad(AES.new(key, AES.MODE_CBC, iv).decrypt(ciphertext), 16)
    must(plaintext == data, "AES roundtrip failed")
    first = get_random_bytes(8)
    second = get_random_bytes(8)
    must(len(first) == 8 and len(second) == 8 and first != second, "Crypto.Random failed")


@check("pycryptodome.hash_hmac")
def _():
    import hashlib
    import hmac
    from Crypto.Hash import HMAC, SHA256

    data = b"singlefile-runtime"
    sha256_hex = SHA256.new(data).hexdigest()
    hmac_hex = HMAC.new(b"secret", data, SHA256).hexdigest()
    must(hashlib.sha256(data).hexdigest() == sha256_hex, "SHA256 mismatch")
    must(hmac.new(b"secret", data, hashlib.sha256).hexdigest() == hmac_hex, "HMAC mismatch")


@check("pycryptodome.rsa")
def _():
    from Crypto.Hash import SHA256
    from Crypto.PublicKey import RSA
    from Crypto.Signature import pkcs1_15

    payload = b"singlefile-runtime"
    key = RSA.generate(1024)
    digest = SHA256.new(payload)
    signature = pkcs1_15.new(key).sign(digest)
    pkcs1_15.new(key.publickey()).verify(SHA256.new(payload), signature)


@check("requests_stack")
def _():
    from collections import OrderedDict
    from pathlib import Path

    import certifi
    import charset_normalizer
    import idna
    import requests
    import urllib3

    prepared = requests.Request(
        "POST",
        "https://example.com/api",
        params=OrderedDict([("a", "1"), ("b", "2")]),
        headers={"X-Test": "singlefile"},
        json={"ok": True},
    ).prepare()
    must(prepared.url == "https://example.com/api?a=1&b=2", "requests URL mismatch")
    must(prepared.headers["Content-Type"].startswith("application/json"), "requests header mismatch")
    url = urllib3.util.parse_url("https://example.com:443/demo?q=1")
    must(url.host == "example.com" and url.port == 443, "urllib3 parse failed")
    cert_path = Path(certifi.where())
    must(cert_path.exists(), "certifi bundle missing")
    must("BEGIN CERTIFICATE" in cert_path.read_text(encoding="utf-8"), "certifi bundle invalid")
    encoding = charset_normalizer.from_bytes("中文".encode("utf-8")).best().encoding
    must("utf" in encoding.lower(), "charset_normalizer failed")
    domain = idna.encode("例子.测试").decode("ascii")
    must(domain.startswith("xn--"), "idna encode failed")
    must(idna.decode(domain) == "例子.测试", "idna decode failed")


@check("async_stack")
def _():
    import anyio
    import sniffio
    from asgiref.sync import async_to_sync

    async def probe():
        return sniffio.current_async_library()

    async def add(a, b):
        return a + b

    must(anyio.run(probe) == "asyncio", "anyio/sniffio failed")
    must(async_to_sync(add)(2, 3) == 5, "asgiref failed")


@check("async_web_stack")
def _():
    import aiohappyeyeballs
    import aiohttp
    import aiosignal
    import frozenlist
    import multidict
    import propcache
    import yarl
    from aiohttp import web
    from websockets.uri import parse_uri

    must(aiohttp.ClientTimeout(total=5).total == 5, "aiohttp timeout failed")

    async def handler(request):
        return web.Response(text="ok")

    app = web.Application()
    route = app.router.add_get("/hello", handler)
    must(route.method == "GET", "aiohttp route failed")

    uri = parse_uri("wss://example.com/chat?q=1")
    must(uri.secure and uri.host == "example.com", "websockets parse_uri failed")

    frozen = frozenlist.FrozenList([1, 2])
    frozen.freeze()
    must(frozen.frozen and list(frozen) == [1, 2], "frozenlist failed")

    values = multidict.MultiDict([("a", "1"), ("a", "2")])
    must(values.getall("a") == ["1", "2"], "multidict failed")
    must(str(yarl.URL("https://example.com") / "demo") == "https://example.com/demo", "yarl failed")
    must(callable(aiohappyeyeballs.start_connection), "aiohappyeyeballs failed")

    class Demo:
        calls = 0

        @propcache.cached_property
        def value(self):
            self.calls += 1
            return self.calls

    demo = Demo()
    must(demo.value == 1 and demo.value == 1, "propcache cached_property failed")

    signal = aiosignal.Signal(owner=None)
    seen = []

    async def receiver(*args, **kwargs):
        seen.append((args, kwargs))

    signal.append(receiver)
    signal.freeze()
    asyncio.run(signal.send("sender", value=1))
    must(seen and seen[0][0] == ("sender",), "aiosignal failed")


@check("attrs_stack")
def _():
    import attr
    import attrs
    import cattr
    import cattrs

    @attr.s(auto_attribs=True)
    class LegacyItem:
        value: int

    @attrs.define
    class ModernItem:
        value: int

    legacy = LegacyItem(3)
    converter = cattrs.Converter()
    structured = converter.structure({"value": 7}, ModernItem)
    must(attr.asdict(legacy)["value"] == 3, "attr failed")
    must(structured.value == 7, "cattrs structure failed")
    must(cattr.unstructure(structured)["value"] == 7, "cattr unstructure failed")


@check("cli_stack")
def _():
    import colorama
    import shellingham
    import six
    import tabulate
    import tenacity
    import tqdm
    import typer
    import typing_extensions
    from click.testing import CliRunner
    import click

    @click.command()
    @click.option("--name")
    def demo(name):
        click.echo(f"hi {name}")

    result = CliRunner().invoke(demo, ["--name", "codex"])
    must(result.exit_code == 0 and "hi codex" in result.output, "click failed")
    must(colorama.ansi.clear_screen() == "\x1b[2J", "colorama failed")
    try:
        shell = shellingham.detect_shell()
    except shellingham.ShellDetectionFailure:
        shell = None
    must(shell is None or (isinstance(shell, tuple) and len(shell) == 2), "shellingham failed")
    must("demo" in tabulate.tabulate([["demo", 1]], headers=["name", "count"]), "tabulate failed")
    must("100%" in tqdm.tqdm.format_meter(1, 1, 1.0), "tqdm failed")
    must(str(tenacity.stop_after_attempt(2)) != "", "tenacity failed")
    must(isinstance(typer.Typer(), typer.main.Typer), "typer failed")
    must(list(six.moves.range(3)) == [0, 1, 2], "six failed")
    annotated = typing_extensions.Annotated[int, "doc"]
    must(
        typing_extensions.get_args(annotated) == (int, "doc"),
        "typing_extensions failed",
    )


@check("text_stack")
def _():
    import markupsafe
    import markdown as markdown_pkg
    import mistune
    import rich.console
    import sqlparse
    import tomlkit
    import wcwidth
    from bs4 import BeautifulSoup
    from markdown_it import MarkdownIt
    from mdurl import decode as mdurl_decode, encode as mdurl_encode, parse as mdurl_parse
    from pygments import highlight
    from pygments.formatters import HtmlFormatter
    from pygments.lexers import PythonLexer

    must(markupsafe.escape("<demo>") == "&lt;demo&gt;", "markupsafe failed")
    soup = BeautifulSoup("<div><p>ok</p></div>", "html.parser")
    must(soup.select_one("p").text == "ok", "bs4/soupsieve failed")
    must("<strong>ok</strong>" in markdown_pkg.markdown("**ok**"), "markdown failed")
    must("<strong>ok</strong>" in mistune.html("**ok**"), "mistune failed")
    must("<strong>ok</strong>" in MarkdownIt().render("**ok**"), "markdown_it failed")
    must(mdurl_decode("a%20b") == "a b", "mdurl decode failed")
    must(mdurl_encode("a b") == "a%20b", "mdurl encode failed")
    parsed = mdurl_parse("https://example.com/demo?q=1")
    must(parsed.hostname == "example.com", "mdurl parse failed")
    html = highlight("print('x')\n", PythonLexer(), HtmlFormatter())
    must("highlight" in html, "pygments failed")
    console_buffer = io.StringIO()
    console = rich.console.Console(file=console_buffer, force_terminal=False, color_system=None)
    console.print("[bold]rich[/bold]")
    must("rich" in console_buffer.getvalue(), "rich failed")
    must(sqlparse.format("select 1", keyword_case="upper").startswith("SELECT"), "sqlparse failed")
    must(tomlkit.parse("answer = 42\n")["answer"] == 42, "tomlkit failed")
    must(wcwidth.wcswidth("中文") == 4, "wcwidth failed")


@check("web_stack")
def _():
    import dateutil.parser
    import dotenv
    import h11
    import httpcore
    import httpx
    import jinja2
    import packaging.version
    import prompt_toolkit.document
    import pymysql
    import redis
    from annotated_doc import Doc

    must(httpx.Request("GET", "https://example.com/demo").url.host == "example.com", "httpx failed")
    must(hasattr(httpcore, "__version__"), "httpcore failed")
    request = h11.Request(method="GET", target="/", headers=[("host", "example.com")])
    must(request.method == b"GET", "h11 failed")
    must(jinja2.Template("Hello {{ name }}").render(name="world") == "Hello world", "jinja2 failed")
    must(packaging.version.parse("1.2.3") < packaging.version.parse("2.0.0"), "packaging failed")
    must(dateutil.parser.isoparse("2024-01-02T03:04:05+00:00").tzinfo is not None, "dateutil failed")
    must(dotenv.dotenv_values(stream=io.StringIO("A=1\n"))["A"] == "1", "dotenv failed")
    document = prompt_toolkit.document.Document("abc", cursor_position=3)
    must(document.cursor_position == 3, "prompt_toolkit failed")
    must(pymysql.converters.escape_string("a'b") == "a\\'b", "pymysql failed")
    pool_kwargs = redis.Redis.from_url("redis://localhost:6379/0").connection_pool.connection_kwargs
    must(pool_kwargs["host"] == "localhost", "redis failed")
    must(repr(Doc("demo")) == "Doc('demo')", "annotated_doc failed")


@check("database_stack")
def _():
    import alembic
    import sqlalchemy as sa
    from alembic.config import Config
    from mako.template import Template

    engine = sa.create_engine("sqlite:///:memory:")
    metadata = sa.MetaData()
    demo = sa.Table(
        "demo",
        metadata,
        sa.Column("id", sa.Integer, primary_key=True),
        sa.Column("name", sa.String),
    )
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(demo.insert().values(name="codex"))
        value = connection.execute(sa.select(demo.c.name)).scalar_one()
    must(value == "codex", "sqlalchemy sqlite roundtrip failed")

    config = Config()
    config.set_main_option("sqlalchemy.url", "sqlite:///:memory:")
    must(config.get_main_option("sqlalchemy.url").startswith("sqlite"), "alembic config failed")
    must(hasattr(alembic, "__version__"), "alembic import failed")
    must(Template("hello ${name}").render(name="codex") == "hello codex", "mako failed")


@check("serialization_stack")
def _():
    import msgpack
    import yaml
    from google.protobuf.struct_pb2 import Struct

    must(yaml.safe_load("answer: 42\n") == {"answer": 42}, "PyYAML safe_load failed")
    packed = msgpack.packb({"items": [1, 2, 3]}, use_bin_type=True)
    must(msgpack.unpackb(packed, raw=False) == {"items": [1, 2, 3]}, "msgpack roundtrip failed")

    payload = Struct()
    payload["answer"] = 42
    payload["name"] = "codex"
    clone = Struct()
    clone.ParseFromString(payload.SerializeToString())
    must(clone["answer"] == 42 and clone["name"] == "codex", "protobuf roundtrip failed")


@check("flask_stack")
def _():
    from blinker import signal
    from flask import Flask
    from itsdangerous import URLSafeSerializer
    from werkzeug.datastructures import MultiDict

    seen = []
    demo_signal = signal("demo")
    demo_signal.connect(lambda sender: seen.append(sender), weak=False)
    demo_signal.send("app")
    must(seen == ["app"], "blinker failed")

    app = Flask(__name__)

    @app.get("/hello")
    def hello():
        return "ok"

    with app.test_client() as client:
        response = client.get("/hello")
        must(response.status_code == 200 and response.data == b"ok", "flask failed")

    serializer = URLSafeSerializer("secret")
    token = serializer.dumps({"x": 1})
    must(serializer.loads(token)["x"] == 1, "itsdangerous failed")
    must(MultiDict([("a", "1"), ("a", "2")]).getlist("a") == ["1", "2"], "werkzeug failed")


@check("django_setup")
def _():
    import django
    from django.conf import settings
    from django.utils import timezone

    if not settings.configured:
        settings.configure(
            SECRET_KEY="singlefile-test",
            INSTALLED_APPS=[],
            USE_TZ=True,
            TIME_ZONE="UTC",
        )
    django.setup()
    must(timezone.is_aware(timezone.now()), "django timezone failed")


@check("math_stack")
def _():
    import mpmath
    import networkx as nx
    import sympy

    must(abs(float(mpmath.sqrt(9)) - 3.0) < 1e-9, "mpmath failed")
    graph = nx.Graph()
    graph.add_edge("a", "b")
    graph.add_edge("b", "c")
    must(nx.shortest_path(graph, "a", "c") == ["a", "b", "c"], "networkx failed")
    x = sympy.Symbol("x")
    must(str(sympy.factor(x**2 - 1)) == "(x - 1)*(x + 1)", "sympy failed")


@check("io_stack")
def _():
    import os
    import et_xmlfile
    import openpyxl
    import portalocker
    import pypdf
    import pyperclip
    import tempfile
    import tzdata
    import win32_setctime
    import xlsxwriter
    from comtypes import GUID

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "demo"
    openpyxl_buffer = io.BytesIO()
    workbook.save(openpyxl_buffer)
    openpyxl_buffer.seek(0)
    loaded_workbook = openpyxl.load_workbook(openpyxl_buffer)
    must(loaded_workbook.active["A1"].value == "demo", "openpyxl failed")

    xlsx_buffer = io.BytesIO()
    xlsx_workbook = xlsxwriter.Workbook(xlsx_buffer, {"in_memory": True})
    xlsx_sheet = xlsx_workbook.add_worksheet("Sheet1")
    xlsx_sheet.write(0, 0, "demo")
    xlsx_workbook.close()
    must(xlsx_buffer.getbuffer().nbytes > 0, "xlsxwriter failed")

    pdf_writer = pypdf.PdfWriter()
    pdf_writer.add_blank_page(width=72, height=72)
    pdf_buffer = io.BytesIO()
    pdf_writer.write(pdf_buffer)
    pdf_buffer.seek(0)
    must(len(pypdf.PdfReader(pdf_buffer).pages) == 1, "pypdf failed")

    must(hasattr(et_xmlfile, "__version__"), "et_xmlfile failed")
    must(isinstance(pyperclip.is_available(), bool), "pyperclip failed")
    must("UTC" in tzdata.available_timezones(), "tzdata failed")
    must(isinstance(win32_setctime.SUPPORTED, bool), "win32_setctime failed")
    must(str(GUID("{00000000-0000-0000-C000-000000000046}")).startswith("{00000000"), "comtypes failed")

    fd, lock_path = tempfile.mkstemp()
    os.close(fd)
    try:
        with portalocker.Lock(lock_path, "w", timeout=1) as locked:
            locked.write("locked")
        with open(lock_path, encoding="utf-8") as lock_file:
            must(lock_file.read() == "locked", "portalocker failed")
    finally:
        os.unlink(lock_path)


@check("loguru_stack")
def _():
    from loguru import logger

    buffer = io.StringIO()
    handler_id = logger.add(buffer, format="{message}")
    try:
        logger.info("demo")
    finally:
        logger.remove(handler_id)
    must("demo" in buffer.getvalue(), "loguru failed")


@check("plotly_stack")
def _():
    import narwhals
    import plotly.graph_objects as go

    figure = go.Figure(data=[go.Bar(x=["a", "b"], y=[1, 2])])
    payload = json.loads(figure.to_json())
    must(payload["data"][0]["type"] == "bar", "plotly figure JSON failed")
    must(callable(narwhals.from_native), "narwhals import failed")


@check("libui_api")
def _():
    import libui

    must(hasattr(libui, "Window"), "libui.Window missing")
    must(hasattr(libui, "Button"), "libui.Button missing")


print(json.dumps(results, ensure_ascii=False, sort_keys=True))
"""


def make_process_failure(step: str, result: dict, name: str | None = None) -> dict:
    return {
        "step": step,
        "name": name or step,
        "error_type": "TimeoutExpired" if result.get("timeout") else "SubprocessError",
        "error": result.get("error") or f"command exited with code {result.get('returncode')}",
        "traceback": "",
        "stdout": result.get("stdout", ""),
        "stderr": result.get("stderr", ""),
        "command": result.get("display"),
    }


def make_json_parse_failure(step: str, result: dict) -> dict:
    return {
        "step": step,
        "name": step,
        "error_type": "JSONDecodeError",
        "error": "verification helper did not return valid JSON",
        "traceback": "",
        "stdout": result.get("stdout", ""),
        "stderr": result.get("stderr", ""),
        "command": result.get("display"),
    }


def summarize_json_failures(step: str, items: list[dict]) -> list[dict]:
    failures = []
    passed = 0
    for item in items:
        if item.get("ok"):
            passed += 1
            continue
        failures.append(
            {
                "step": step,
                "name": item.get("name", step),
                "error_type": item.get("error_type", "Error"),
                "error": item.get("error", "unknown error"),
                "traceback": item.get("traceback", ""),
                "stdout": "",
                "stderr": "",
                "command": "",
            }
        )
    log(f"{step}: {passed} passed, {len(failures)} failed")
    return failures


def run_json_step(step: str, python_exe: Path, code: str, cwd: Path, timeout: float) -> list[dict]:
    result = run_capture([str(python_exe), "-c", code], cwd=cwd, timeout=timeout)
    if result.get("timeout") or not result.get("ok"):
        return [make_process_failure(step, result)]
    try:
        items = json.loads(result.get("stdout", ""))
    except json.JSONDecodeError:
        return [make_json_parse_failure(step, result)]
    return summarize_json_failures(step, items)


def run_command_step(step: str, cmd: list[str], cwd: Path, timeout: float) -> list[dict]:
    result = run_capture(cmd, cwd=cwd, timeout=timeout)
    if result.get("ok"):
        log(f"{step}: passed")
        return []
    return [make_process_failure(step, result)]


def verify_imports_and_smoke(
    python_exe: Path,
    source_root: Path,
    manifest: dict,
    integrations: list,
    cwd: Path,
    run_third_party_smoke: bool,
) -> list[dict]:
    failures = []
    failures.extend(
        run_json_step(
            "imports",
            python_exe,
            build_import_check_code(source_root, manifest, integrations),
            cwd,
            300,
        )
    )
    failures.extend(run_json_step("stdlib-smoke", python_exe, STDLIB_SMOKE, cwd, 300))
    if integrations and run_third_party_smoke:
        failures.extend(run_json_step("third-party-smoke", python_exe, THIRD_PARTY_SMOKE, cwd, 300))
    elif integrations:
        log("skip third-party-smoke because this profile selects an incremental library subset")
    else:
        log("skip third-party-smoke because no third-party integrations are enabled")
    return failures


def verify_integration_steps(
    python_exe: Path,
    repo_root: Path,
    integrations: list,
    skipped_groups: set[str],
) -> list[dict]:
    failures: list[dict] = []
    for step in collect_verification_steps(integrations):
        skip_group = step.get("skip_group")
        if skip_group and skip_group in skipped_groups:
            log(f"skip verification step {step['name']} because --skip-{skip_group} was requested")
            continue

        timeout = float(step.get("timeout", 240))
        if step["kind"] == "module":
            failures.extend(
                run_command_step(
                    step["name"],
                    [str(python_exe), "-m", step["module"]],
                    repo_root,
                    timeout,
                )
            )
            continue
        if step["kind"] == "script":
            failures.extend(
                run_command_step(
                    step["name"],
                    [str(python_exe), str(repo_root / step["script"])],
                    repo_root,
                    timeout,
                )
            )
            continue
        if step["kind"] == "inline":
            failures.extend(
                run_command_step(
                    step["name"],
                    [str(python_exe), "-c", step["code"]],
                    repo_root,
                    timeout,
                )
            )
            continue
        raise RuntimeError(f"unsupported verification step kind: {step['kind']}")
    return failures


def verification_coverage(integrations: list) -> dict:
    explicit_steps = collect_verification_steps(integrations)
    libraries_without_steps = sorted(
        integration.name for integration in integrations if not integration.verification_steps
    )
    libraries_with_steps = sorted(
        integration.name for integration in integrations if integration.verification_steps
    )
    return {
        "library_count": len(integrations),
        "explicit_step_count": len(explicit_steps),
        "libraries_with_steps": libraries_with_steps,
        "libraries_without_steps": libraries_without_steps,
    }


def emit_failure(failure: dict, index: int, total: int) -> None:
    header = f"[{index}/{total}] {failure['step']}::{failure['name']}"
    log(header)
    log(f"  {failure.get('error_type', 'Error')}: {failure.get('error', 'unknown error')}")
    command = failure.get("command")
    if command:
        log(f"  command: {command}")
    traceback_text = failure.get("traceback", "").strip()
    if traceback_text:
        log("  traceback:")
        for line in traceback_text.splitlines():
            log(f"    {line}")
    stdout_text = failure.get("stdout", "").strip()
    if stdout_text:
        log("  stdout:")
        for line in stdout_text.splitlines():
            log(f"    {line}")
    stderr_text = failure.get("stderr", "").strip()
    if stderr_text:
        log("  stderr:")
        for line in stderr_text.splitlines():
            log(f"    {line}")


def write_report(path: Path, python_exe: Path, failures: list[dict], coverage: dict) -> None:
    report = {
        "python_exe": str(python_exe),
        "failure_count": len(failures),
        "failures": failures,
        "verification_coverage": coverage,
    }
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"wrote verification report to {path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify a single-file Python runtime and bundled libraries.")
    parser.add_argument("--python-exe", type=Path, required=True, help="Path to the built single-file python.exe")
    parser.add_argument("--manifest", type=Path, required=True, help="Path to manifest.json")
    parser.add_argument("--config", type=Path, help="Path to config.json")
    parser.add_argument("--repo-root", type=Path, required=True, help="Path to the builder repository root")
    parser.add_argument("--source-root", type=Path, required=True, help="Path to the patched CPython source tree")
    parser.add_argument("--profile", help="Build profile from config.json. Defaults to config.default_profile.")
    parser.add_argument("--skip-crypto", action="store_true", help="Skip the full Crypto.SelfTest suite")
    parser.add_argument("--skip-gui", action="store_true", help="Skip libui smoke and GUI tests")
    parser.add_argument("--report-json", type=Path, help="Optional path to write a JSON verification report")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    python_exe = args.python_exe.resolve()
    manifest = load_manifest(args.manifest.resolve())
    repo_root = args.repo_root.resolve()
    config = load_config((args.config or (repo_root / "config.json")).resolve())
    profile_name, profile = resolve_profile(config, args.profile)
    integrations = load_integrations(repo_root / "Lib", profile.get("third_party_libraries", "all"))
    log(f"verification profile: {profile_name} ({len(integrations)} third-party integration(s))")
    coverage = verification_coverage(integrations)
    missing_steps = coverage["libraries_without_steps"]
    log(
        "explicit per-library verification: "
        f"{len(coverage['libraries_with_steps'])}/{coverage['library_count']} libraries, "
        f"{coverage['explicit_step_count']} step(s)"
    )
    if missing_steps:
        preview = ", ".join(missing_steps[:20])
        suffix = f" ... (+{len(missing_steps) - 20} more)" if len(missing_steps) > 20 else ""
        log(f"verification gaps still import-only: {preview}{suffix}")
    source_root = args.source_root.resolve()
    if not python_exe.exists():
        raise RuntimeError(f"python executable not found: {python_exe}")

    skipped_groups = set()
    if args.skip_crypto:
        skipped_groups.add("crypto")
    if args.skip_gui:
        skipped_groups.add("gui")

    failures = []
    run_third_party_smoke = profile.get("third_party_libraries") == "all"
    failures.extend(
        verify_imports_and_smoke(
            python_exe,
            source_root,
            manifest,
            integrations,
            repo_root,
            run_third_party_smoke,
        )
    )
    failures.extend(verify_integration_steps(python_exe, repo_root, integrations, skipped_groups))

    if args.report_json:
        write_report(args.report_json.resolve(), python_exe, failures, coverage)

    if failures:
        log(f"verification failed with {len(failures)} issue(s)")
        for index, failure in enumerate(failures, start=1):
            emit_failure(failure, index, len(failures))
        raise SystemExit(1)

    log("all verification steps passed")


if __name__ == "__main__":
    main()
