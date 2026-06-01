from __future__ import annotations

import traceback
import os
import subprocess
import sys
import tempfile
import types
from pathlib import Path


SMOKE_TESTS = [
    (
        'aiohappyeyeballs-smoke',
        r"""
import inspect
import aiohappyeyeballs
from aiohappyeyeballs import addr_to_addr_infos

infos = addr_to_addr_infos(("127.0.0.1", 80))
assert infos and infos[0][4] == ("127.0.0.1", 80)
assert inspect.iscoroutinefunction(aiohappyeyeballs.start_connection)
        """,
    ),
    (
        'aiohttp-smoke',
        r"""
import importlib.util

import aiohttp
import aiohttp._http_parser as http_parser_ext
import aiohttp._http_writer as http_writer_ext
import aiohttp._websocket.mask as mask_ext
import aiohttp._websocket.reader_c as reader_ext
from multidict import CIMultiDict

for name in (
    "aiohttp._http_parser",
    "aiohttp._http_writer",
    "aiohttp._websocket.mask",
    "aiohttp._websocket.reader_c",
):
    assert importlib.util.find_spec(name).origin == "built-in", name

headers = CIMultiDict([("Host", "example.com"), ("X-Test", "ok")])
serialized = http_writer_ext._serialize_headers("GET / HTTP/1.1", headers)
assert serialized == b"GET / HTTP/1.1\r\nHost: example.com\r\nX-Test: ok\r\n\r\n", serialized

data = bytearray(b"abcd")
mask_ext._websocket_mask_cython(b"\x01\x02\x03\x04", data)
assert data == bytearray([0x60, 0x60, 0x60, 0x60]), data

request = http_parser_ext.RawRequestMessage(
    "GET",
    "/",
    aiohttp.HttpVersion11,
    headers,
    tuple(headers.items()),
    False,
    None,
    False,
    False,
    None,
)
assert request.method == "GET" and request.path == "/", request
assert hasattr(reader_ext, "WebSocketReader")
        """,
    ),
    (
        'aiosignal-smoke',
        r"""
import asyncio
from aiosignal import Signal

seen = []
signal = Signal(owner="owner")

async def receiver(*args, **kwargs):
    seen.append((args, kwargs))

signal.append(receiver)
signal.freeze()
asyncio.run(signal.send("sender", value=3))
assert seen == [(('sender',), {'value': 3})]
try:
    signal.append(receiver)
except RuntimeError:
    pass
else:
    raise AssertionError("aiosignal allowed mutation after freeze")
        """,
    ),
    (
        'alembic-smoke',
        r"""
from alembic.config import Config
from alembic.script.revision import RevisionMap, Revision

config = Config()
config.set_main_option("sqlalchemy.url", "sqlite:///:memory:")
assert config.get_main_option("sqlalchemy.url") == "sqlite:///:memory:"
rev = Revision("abc", None)
rev_map = RevisionMap(lambda: [rev])
assert rev_map.get_revision("abc").revision == "abc"
        """,
    ),
    (
        'annotated-doc-smoke',
        r"""
from annotated_doc import Doc

marker = Doc("primary key")
assert marker.documentation == "primary key"
assert repr(marker) == "Doc('primary key')"
        """,
    ),
    (
        'annotated-types-smoke',
        r"""
from typing import Annotated

from annotated_types import Ge, Len

field = Annotated[str, Len(min_length=2), Ge("aa")]
metadata = field.__metadata__
assert isinstance(metadata[0], Len)
assert metadata[0].min_length == 2
assert metadata[1].ge == "aa"
        """,
    ),
    (
        'anyio-smoke',
        r"""
import anyio

async def worker():
    send, receive = anyio.create_memory_object_stream(1)
    async with send, receive:
        await send.send("ok")
        value = await receive.receive()

    seen = []

    async def child():
        seen.append(anyio.get_current_task().name is not None)

    async with anyio.create_task_group() as group:
        group.start_soon(child)

    return value, seen

assert anyio.run(worker) == ("ok", [True])
        """,
    ),
    (
        'asgiref-smoke',
        r"""
import asyncio
from asgiref.sync import async_to_sync, sync_to_async

async def add(a, b):
    return a + b

def mul(a, b):
    return a * b

assert async_to_sync(add)(2, 3) == 5
assert asyncio.run(sync_to_async(mul)(4, 5)) == 20
        """,
    ),
    (
        'asttokens-smoke',
        r"""
import asttokens

tokens = asttokens.ASTTokens("x = 1 + 2", parse=True)
assign = tokens.tree.body[0]
assert tokens.get_text(assign.value) == "1 + 2"
start, end = tokens.get_text_range(assign.targets[0])
assert tokens.text[start:end] == "x"
        """,
    ),
    (
        'async-lru-smoke',
        r"""
import asyncio

from async_lru import alru_cache


calls = {"count": 0}


@alru_cache(maxsize=4)
async def cached(value):
    calls["count"] += 1
    await asyncio.sleep(0)
    return value * 2


async def main():
    first = await cached(21)
    second = await cached(21)
    info = cached.cache_info()
    assert first == 42
    assert second == 42
    assert calls["count"] == 1
    assert info.hits >= 1
    assert cached.cache_contains(21)
    assert cached.cache_invalidate(21)
    assert not cached.cache_contains(21)
    await cached.cache_close()


asyncio.run(main())
        """,
    ),
    (
        'attr-smoke',
        r"""
import attr

@attr.s(auto_attribs=True, slots=True)
class Item:
    value: int = attr.ib(validator=attr.validators.instance_of(int))

item = Item(5)
assert attr.asdict(item) == {"value": 5}
assert attr.evolve(item, value=6).value == 6
        """,
    ),
    (
        'attrs-smoke',
        r"""
import attrs

@attrs.define(frozen=True)
class Point:
    x: int = attrs.field(validator=attrs.validators.ge(0))
    y: int = 0

point = Point(1, 2)
assert attrs.asdict(point) == {"x": 1, "y": 2}
assert attrs.evolve(point, y=3).y == 3
        """,
    ),
    (
        'babel-smoke',
        r"""
from datetime import datetime
from babel import Locale
from babel.dates import format_datetime
from babel.localedata import exists
from babel.numbers import format_currency

locale = Locale.parse("en_US")
formatted_datetime = format_datetime(datetime(2024, 1, 2, 3, 4, 5), locale="en_US")
formatted_currency = format_currency(1234.5, "USD", locale="en_US")

assert locale.display_name == "English (United States)"
assert "2024" in formatted_datetime
assert "3:04:05" in formatted_datetime
assert "$1,234.50" == formatted_currency
assert exists("root")
assert exists("en")
assert exists("en_US")
assert exists("fr_FR")
        """,
    ),
    (
        'black-smoke',
        r"""
import black
import black.schema

source = "def add(a,b):\n return a+b\n"
formatted = black.format_str(source, mode=black.FileMode())
assert "def add(a, b):" in formatted
assert "return a + b" in formatted
schema = black.schema.get_schema()
assert schema["type"] == "object"
assert "line-length" in schema["properties"]
        """,
    ),
    (
        'blackd-smoke',
        r"""
import blackd

app = blackd.make_app()
assert app is not None
assert len(app.router.routes()) == 1
        """,
    ),
    (
        'bleach-smoke',
        r"""
import bleach
from bleach.css_sanitizer import CSSSanitizer

cleaned = bleach.clean(
    '<p style="color:red;position:absolute"><script>x</script>Hello <b>world</b></p>',
    tags={"p", "b"},
    attributes={"p": ["style"]},
    css_sanitizer=CSSSanitizer(allowed_css_properties=["color"]),
    strip=True,
)
assert "<script>" not in cleaned
assert "position:absolute" not in cleaned
assert "color:red" in cleaned
assert "<b>world</b>" in cleaned
        """,
    ),
    (
        'blinker-smoke',
        r"""
from blinker import Namespace, signal

seen = []
signal("global-demo").connect(lambda sender, **kw: seen.append((sender, kw)), weak=False)
signal("global-demo").send("sender", value=7)
assert seen == [("sender", {"value": 7})]
local = Namespace().signal("local")
assert local.name == "local"
        """,
    ),
    (
        'bs4-smoke',
        r"""
from bs4 import BeautifulSoup

soup = BeautifulSoup("<html><body><p class='x'>one</p><p>two</p></body></html>", "html.parser")
assert soup.select_one("p.x").text == "one"
assert [p.text for p in soup.find_all("p")] == ["one", "two"]
        """,
    ),
    (
        'cachetools-smoke',
        r"""
from cachetools import LRUCache, cached

cache = LRUCache(maxsize=2)
cache["a"] = 1
cache["b"] = 2
cache["c"] = 3
assert "a" not in cache and cache["b"] == 2 and cache["c"] == 3

calls = {"count": 0}

@cached(cache={})
def add(a, b):
    calls["count"] += 1
    return a + b

assert add(2, 3) == 5
assert add(2, 3) == 5
assert calls["count"] == 1
        """,
    ),
    (
        'cattr-smoke',
        r"""
import attr
import cattr

@attr.s(auto_attribs=True)
class Item:
    value: int

item = cattr.structure({"value": 9}, Item)
assert item.value == 9
assert cattr.unstructure(item) == {"value": 9}
        """,
    ),
    (
        'cattrs-smoke',
        r"""
import attrs
import cattrs

@attrs.define
class Item:
    value: int

converter = cattrs.Converter()
item = converter.structure({"value": 11}, Item)
assert item.value == 11
assert converter.unstructure(item) == {"value": 11}
        """,
    ),
    (
        'certifi-smoke',
        r"""
from pathlib import Path
import certifi

path = Path(certifi.where())
contents = certifi.contents()
assert path.exists()
assert "BEGIN CERTIFICATE" in contents
assert path.read_text(encoding="ascii") == contents
        """,
    ),
    (
        'chardet-smoke',
        r"""
import chardet
from chardet.models import BigramProfile, get_enc_index, get_idf_weights, load_models
from chardet.pipeline.confusion import load_confusion_data
from chardet.universaldetector import UniversalDetector

payload = b"caf\xe9"
result = chardet.detect(payload)
assert result["encoding"].lower() in {"iso-8859-1", "windows-1255", "windows-1252"}

detector = UniversalDetector()
detector.feed(payload)
detector.close()
assert detector.result["encoding"]

models = load_models()
enc_index = get_enc_index()
idf = get_idf_weights()
confusion = load_confusion_data()
profile = BigramProfile(payload)
assert models
assert enc_index
assert len(idf) == 65536
assert confusion
assert profile.nonzero
        """,
    ),
    (
        'charset-normalizer-smoke',
        r"""
from charset_normalizer import from_bytes

result = from_bytes(b"caf\xc3\xa9").best()
assert result is not None
assert "utf" in result.encoding.lower()
assert str(result) == "cafe" or str(result) == "café"
        """,
    ),
    (
        'click-smoke',
        r"""
import click
from click.testing import CliRunner

@click.group()
def cli():
    pass

@cli.command()
@click.argument("name")
def hello(name):
    click.echo(f"hello {name}")

result = CliRunner().invoke(cli, ["hello", "codex"])
assert result.exit_code == 0
assert "hello codex" in result.output
        """,
    ),
    (
        'cloudpickle-smoke',
        r"""
import cloudpickle

factor = 40

def make_adder(delta):
    return lambda value: value + delta + factor

class LocalGreeter:
    def __init__(self, prefix):
        self.prefix = prefix

    def greet(self, name):
        return f"{self.prefix} {name}"

func = cloudpickle.loads(cloudpickle.dumps(make_adder(1)))
greeter = cloudpickle.loads(cloudpickle.dumps(LocalGreeter("hi")))
assert func(1) == 42
assert greeter.greet("codex") == "hi codex"
        """,
    ),
    (
        'colorama-smoke',
        r"""
from colorama import Fore, Style, ansi, just_fix_windows_console

just_fix_windows_console()
assert Fore.RED.startswith("[")
assert Style.RESET_ALL.endswith("m")
assert ansi.clear_screen() == "[2J"
        """,
    ),
    (
        'comm-smoke',
        r"""
from comm import BaseComm, create_comm, get_comm_manager


class RecordingComm(BaseComm):
    def __init__(self, *args, **kwargs):
        self.records = []
        super().__init__(*args, **kwargs)

    def publish_msg(self, msg_type, data=None, metadata=None, buffers=None, **keys):
        self.records.append((msg_type, data or {}, metadata or {}, list(buffers or []), keys))


manager = get_comm_manager()
manager.targets.clear()
manager.comms.clear()

default_comm = create_comm(target_name="default", comm_id="default-comm", data={"seed": 1})
assert manager.get_comm("default-comm") is default_comm
default_comm.close({"done": False})
assert manager.get_comm("default-comm") is None

opened = []
manager.register_target("staticpython-test", lambda comm, msg: opened.append((comm.comm_id, msg["content"]["data"])))

comm = RecordingComm(
    target_name="staticpython-test",
    comm_id="staticpython-comm",
    data={"hello": "world"},
)
assert manager.get_comm("staticpython-comm") is comm
manager.targets["staticpython-test"](comm, {"content": {"data": {"hello": "world"}}})

received = []
closed = []
comm.on_msg(lambda msg: received.append(msg["content"]["data"]))
comm.on_close(lambda msg: closed.append(msg["content"]["data"]))
comm.handle_msg({"content": {"data": {"answer": 42}}})
comm.handle_close({"content": {"data": {"done": True}}})

comm.send({"answer": 42})
comm.close({"done": True})

assert isinstance(comm, RecordingComm)
assert opened == [("staticpython-comm", {"hello": "world"})]
assert received == [{"answer": 42}]
assert closed == [{"done": True}]
assert [item[0] for item in comm.records] == ["comm_open", "comm_msg", "comm_close"]
assert comm.records[0][1]["hello"] == "world"
assert comm.records[1][1]["answer"] == 42
        """,
    ),
    (
        'comtypes-smoke',
        r"""
from comtypes import GUID, HRESULT, COMMETHOD

guid = GUID("{00000000-0000-0000-C000-000000000046}")
assert str(guid).startswith("{00000000")
assert HRESULT(0).value == 0
method = COMMETHOD([], HRESULT, "Demo")
assert method.name == "Demo"
assert method.restype is HRESULT
        """,
    ),
    (
        'cycler-smoke',
        r"""
from cycler import cycler

combined = cycler(color=["red", "blue"]) + cycler(linewidth=[1, 2])
assert list(combined) == [
    {"color": "red", "linewidth": 1},
    {"color": "blue", "linewidth": 2},
]
product = cycler(color=["red", "blue"]) * cycler(marker=["o", "x"])
assert len(product) == 4
assert {"color": "red", "marker": "o"} in list(product)
        """,
    ),
    (
        'contourpy-smoke',
        r"""
import importlib.util

import numpy as np

import contourpy
from contourpy.util import build_config

assert importlib.util.find_spec("contourpy._contourpy").origin == "built-in"
assert contourpy.__version__

z = np.array(
    [
        [0.0, 0.5, 1.0],
        [0.5, 1.0, 1.5],
        [1.0, 1.5, 2.0],
    ],
    dtype=np.float64,
)
generator = contourpy.contour_generator(z=z, name="serial")
lines = generator.lines(0.75)
assert lines
assert all(line.shape[1] == 2 for line in lines)
assert build_config()["contourpy_version"] == contourpy.__version__
        """,
    ),
    (
        'dateutil-smoke',
        r"""
from datetime import datetime
from dateutil import parser, tz
from dateutil.relativedelta import relativedelta

value = parser.isoparse("2024-01-02T03:04:05+00:00")
assert value.tzinfo is not None
assert datetime(2024, 1, 31) + relativedelta(months=1) == datetime(2024, 2, 29)
assert tz.gettz("UTC") is not None
        """,
    ),
    (
        'decorator-smoke',
        r"""
from decorator import decorator

@decorator
def traced(func, *args, **kwargs):
    return ("called", func(*args, **kwargs))

@traced
def add(a, b):
    return a + b

assert add(2, 4) == ("called", 6)
assert add.__name__ == "add"
        """,
    ),
    (
        'defusedxml-smoke',
        r"""
from defusedxml import ElementTree

root = ElementTree.fromstring("<root><child value='1'/></root>")
assert root.tag == "root"
assert root[0].attrib["value"] == "1"
        """,
    ),
    (
        'dill-smoke',
        r"""
import dill
import sys

state = {"base": 40}

def add_from_state(value):
    return state["base"] + value

payload = dill.loads(dill.dumps({"answer": 42}))
assert payload["answer"] == 42
if sys.version_info < (3, 15):
    func = dill.loads(dill.dumps(lambda value: value * 3))
    stateful = dill.loads(dill.dumps(add_from_state))
    assert func(14) == 42
    assert stateful(2) == 42
        """,
    ),
    (
        'django-smoke',
        r"""
import django
from django.conf import settings
from django.db import connection
from django.http import HttpRequest, HttpResponse, QueryDict
from django.template import Context, Engine
from django.urls import path, resolve, reverse
from django.utils import timezone

if not settings.configured:
    settings.configure(
        SECRET_KEY="staticpython",
        INSTALLED_APPS=[],
        ROOT_URLCONF=__name__,
        USE_TZ=True,
        TIME_ZONE="UTC",
        DATABASES={"default": {"ENGINE": "django.db.backends.sqlite3", "NAME": ":memory:"}},
        TEMPLATES=[{"BACKEND": "django.template.backends.django.DjangoTemplates", "DIRS": [], "APP_DIRS": False, "OPTIONS": {}}],
    )
django.setup()

def index(request):
    return HttpResponse("ok")

urlpatterns = [path("demo/", index, name="demo")]
match = resolve("/demo/")
assert match.func is index
assert reverse("demo") == "/demo/"
request = HttpRequest()
request.method = "GET"
request.path = "/demo/"
request.GET = QueryDict("x=1")
assert index(request).content == b"ok"
with connection.cursor() as cursor:
    cursor.execute("select 1")
    assert cursor.fetchone() == (1,)
assert Engine().from_string("Hello {{ name }}").render(Context({"name": "Codex"})) == "Hello Codex"
assert timezone.is_aware(timezone.now())
        """,
    ),
    (
        'dnspython-smoke',
        r"""
import dns.name
import dns.rdatatype
from dns.message import make_query

name = dns.name.from_text("example.com.")
assert name.to_text() == "example.com."
message = make_query(name, dns.rdatatype.A)
assert message.question[0].name == name
        """,
    ),
    (
        'docutils-smoke',
        r"""
from docutils.core import publish_parts

source = "Title\n=====\n\n* item one\n* item two\n"
parts = publish_parts(source, writer_name="html5")
html = parts["html_body"]
whole = parts["whole"]
assert "<h1" in html and "Title" in html
assert "item one" in html and "item two" in html
assert "<style" in whole and "minimal.css" in whole
        """,
    ),
    (
        'dotenv-smoke',
        r"""
import io
from dotenv import dotenv_values

values = dotenv_values(stream=io.StringIO("A=1\nQUOTED='two words'\n"))
assert values["A"] == "1"
assert values["QUOTED"] == "two words"
        """,
    ),
    (
        'email-validator-smoke',
        r"""
from email_validator import EmailNotValidError, validate_email

result = validate_email("User.Name+tag@example.com", check_deliverability=False)
assert result.normalized == "User.Name+tag@example.com"

try:
    validate_email("not an address", check_deliverability=False)
except EmailNotValidError:
    pass
else:
    raise AssertionError("invalid email was accepted")
        """,
    ),
    (
        'et-xmlfile-smoke',
        r"""
import io
from et_xmlfile import xmlfile

buffer = io.BytesIO()
with xmlfile(buffer) as xf:
    with xf.element("root", {"kind": "demo"}):
        with xf.element("child"):
            xf.write("demo")
xml = buffer.getvalue()
assert b'<root kind="demo">' in xml
assert b"<child>demo</child>" in xml
        """,
    ),
    (
        'executing-smoke',
        r"""
import linecache
import sys
import textwrap
from executing import Source, only

filename = "<executing_smoke>"
source_text = textwrap.dedent(
    '''
    import sys
    def inspect_frame(frame):
        return Source.executing(frame)
    def probe():
        value = 21
        return inspect_frame(sys._getframe())
    '''
)
linecache.cache[filename] = (len(source_text), None, source_text.splitlines(True), filename)
namespace = {"Source": Source, "sys": sys}
exec(compile(source_text, filename, "exec"), namespace)
execution = namespace["probe"]()
assert type(execution.node).__name__ == "Call"
assert execution.text() == "inspect_frame(sys._getframe())"
assert only([42]) == 42
        """,
    ),
    (
        'faker-smoke',
        r"""
from faker import Faker

Faker.seed(12345)
fake = Faker("en_US")
name = fake.name()
email = fake.email()
profile = fake.simple_profile()
assert isinstance(name, str) and " " in name
assert "@" in email
assert {"username", "name", "sex", "address", "mail", "birthdate"} <= set(profile)
        """,
    ),
    (
        'fastjsonschema-smoke',
        r"""
import fastjsonschema

validate = fastjsonschema.compile({
    "type": "object",
    "properties": {"answer": {"type": "integer"}},
    "required": ["answer"],
})
assert validate({"answer": 42}) == {"answer": 42}
try:
    validate({"answer": "nope"})
except fastjsonschema.JsonSchemaException:
    pass
else:
    raise AssertionError("fastjsonschema accepted an invalid payload")
        """,
    ),
    (
        'filelock-smoke',
        r"""
import tempfile
from pathlib import Path

from filelock import FileLock, Timeout

with tempfile.TemporaryDirectory() as temp_dir:
    lock_path = Path(temp_dir) / "demo.lock"
    with FileLock(str(lock_path), timeout=1):
        assert lock_path.exists()
        try:
            FileLock(str(lock_path), timeout=0).acquire()
        except Timeout:
            pass
        else:
            raise AssertionError("filelock allowed a second exclusive lock")
        """,
    ),
    (
        'flask-smoke',
        r"""
from flask import Blueprint, Flask, jsonify, render_template_string, request, url_for

app = Flask(__name__)
api = Blueprint("api", __name__)

@app.get("/hello/<name>")
def hello(name):
    return jsonify(name=name, query=request.args.get("q"))

@api.post("/sum")
def total():
    payload = request.get_json()
    return jsonify(total=sum(payload["values"]))

app.register_blueprint(api, url_prefix="/api")

with app.test_client() as client:
    response = client.get("/hello/codex?q=ok")
    assert response.status_code == 200
    assert response.get_json() == {"name": "codex", "query": "ok"}
    response = client.post("/api/sum", json={"values": [1, 2, 3]})
    assert response.status_code == 200
    assert response.get_json() == {"total": 6}
    with app.test_request_context():
        assert url_for("hello", name="x") == "/hello/x"
    with app.app_context():
        assert render_template_string("Hello {{ name }}", name="Codex") == "Hello Codex"
        """,
    ),
    (
        'fonttools-smoke',
        r"""
from fontTools.misc.transform import Transform
from fontTools.ttLib import TTFont, newTable

transform = Transform().scale(2, 3).translate(5, 7)
assert transform.transformPoint((1, 1)) == (12, 24)
font = TTFont(recalcBBoxes=False, recalcTimestamp=False)
font["name"] = newTable("name")
assert "name" in font
        """,
    ),
    (
        'freezegun-smoke',
        r"""
from datetime import date, datetime, timedelta
from freezegun import freeze_time


@freeze_time("2024-01-02 03:04:05")
def run():
    assert datetime.now() == datetime(2024, 1, 2, 3, 4, 5)
    assert date.today() == date(2024, 1, 2)
    assert datetime.now() + timedelta(days=1) == datetime(2024, 1, 3, 3, 4, 5)


run()
        """,
    ),
    (
        'frozenlist-smoke',
        r"""
from frozenlist import FrozenList

items = FrozenList([1])
items.append(2)
items.freeze()
assert list(items) == [1, 2]
assert items.frozen
try:
    items.append(3)
except RuntimeError:
    pass
else:
    raise AssertionError("FrozenList allowed mutation after freeze")
        """,
    ),
    (
        'fsspec-smoke',
        r"""
from fsspec.core import url_to_fs
from fsspec.implementations.memory import MemoryFileSystem

fs = MemoryFileSystem()
with fs.open("/demo.txt", "wb") as handle:
    handle.write(b"staticpython")
with fs.open("/demo.txt", "rb") as handle:
    assert handle.read() == b"staticpython"

local_fs, path = url_to_fs("file:///tmp/staticpython.txt")
assert path.endswith("staticpython.txt")
assert local_fs.protocol in ("file", ("file", "local"))
        """,
    ),
    (
        'h11-smoke',
        r"""
import h11

conn = h11.Connection(h11.CLIENT)
data = conn.send(h11.Request(method="GET", target="/", headers=[("host", "example.com")]))
assert b"GET / HTTP/1.1" in data
assert conn.our_state is h11.SEND_BODY
conn.send(h11.EndOfMessage())
assert conn.our_state is h11.DONE
        """,
    ),
    (
        'h2-smoke',
        r"""
from h2.config import H2Configuration
from h2.connection import H2Connection
from h2.events import RemoteSettingsChanged

conn = H2Connection(config=H2Configuration(client_side=True, header_encoding="utf-8"))
conn.initiate_connection()
data = conn.data_to_send()
assert data.startswith(b"PRI * HTTP/2.0")

server = H2Connection(config=H2Configuration(client_side=False, header_encoding="utf-8"))
events = server.receive_data(data)
assert any(isinstance(event, RemoteSettingsChanged) for event in events)
server.initiate_connection()
assert server.data_to_send()
        """,
    ),
    (
        'hpack-smoke',
        r"""
from hpack import Decoder, Encoder

headers = [(b":method", b"GET"), (b":path", b"/")]
encoder = Encoder()
encoded = encoder.encode(headers)
decoder = Decoder()
assert decoder.decode(encoded) == [(":method", "GET"), (":path", "/")]
encoded_response = encoder.encode([(":status", "200"), ("content-type", "text/plain")])
assert decoder.decode(encoded_response) == [(":status", "200"), ("content-type", "text/plain")]
        """,
    ),
    (
        'html5lib-smoke',
        r"""
import html5lib

document = html5lib.parse("<!doctype html><title>StaticPython</title><p>ok</p>")
html = document.find("{http://www.w3.org/1999/xhtml}head")
title = html.find("{http://www.w3.org/1999/xhtml}title")
assert title.text == "StaticPython"
fragment = html5lib.parseFragment("<p>one<b>two</p>")
paragraph = fragment.find("{http://www.w3.org/1999/xhtml}p")
assert paragraph is not None
assert paragraph.find("{http://www.w3.org/1999/xhtml}b").text == "two"
        """,
    ),
    (
        'httpcore-smoke',
        r"""
import httpcore

origin = httpcore.Origin(b"https", b"example.com", 443)
assert origin.scheme == b"https"
assert origin.host == b"example.com"
assert origin.port == 443
assert issubclass(httpcore.ConnectError, httpcore.NetworkError)
        """,
    ),
    (
        'httpx-smoke',
        r"""
import asyncio
import httpx

transport = httpx.MockTransport(lambda request: httpx.Response(200, json={"path": request.url.path}))
with httpx.Client(transport=transport, base_url="https://example.com") as client:
    response = client.get("/demo")
assert response.status_code == 200
assert response.json() == {"path": "/demo"}

async def main():
    async def handler(request):
        assert request.url.params["q"] == "x"
        return httpx.Response(201, headers={"X-Test": "ok"}, json={"path": request.url.path})

    transport = httpx.MockTransport(handler)
    async with httpx.AsyncClient(transport=transport, base_url="https://example.com") as client:
        response = await client.get("/items", params={"q": "x"})
        assert response.status_code == 201
        assert response.headers["X-Test"] == "ok"
        assert response.json() == {"path": "/items"}

asyncio.run(main())
        """,
    ),
    (
        'humanize-smoke',
        r"""
from datetime import timedelta
import humanize

assert humanize.intcomma(1234567) == "1,234,567"
assert humanize.naturalsize(1536) == "1.5 kB"
assert "hour" in humanize.naturaldelta(timedelta(hours=2, minutes=5))
        """,
    ),
    (
        'hyperframe-smoke',
        r"""
from hyperframe.frame import DataFrame, SettingsFrame

frame = SettingsFrame(stream_id=0)
serialized = frame.serialize()
parsed, length = SettingsFrame.parse_frame_header(memoryview(serialized[:9]))
parsed.parse_body(memoryview(serialized[9:9 + length]))
assert parsed.stream_id == 0

data_frame = DataFrame(stream_id=1)
data_frame.data = b"hello"
data_frame.flags.add("END_STREAM")
payload = data_frame.serialize()
parsed_data, data_length = DataFrame.parse_frame_header(memoryview(payload[:9]))
parsed_data.parse_body(memoryview(payload[9:9 + data_length]))
assert parsed_data.data == b"hello"
assert "END_STREAM" in parsed_data.flags
        """,
    ),
    (
        'hypothesis-smoke',
        r"""
from hypothesis import given, settings
from hypothesis import strategies as st

seen = []

@settings(max_examples=5, derandomize=True)
@given(st.integers(min_value=0, max_value=10))
def check(value):
    seen.append(value)
    assert value >= 0

check()
assert seen
        """,
    ),
    (
        'idna-smoke',
        r"""
import idna

unicode_domain = "例子.测试"
punycode = idna.encode(unicode_domain).decode("ascii")
assert punycode == "xn--fsqu00a.xn--0zwm56d"
assert idna.decode(punycode) == unicode_domain
assert idna.encode("example.com").decode("ascii") == "example.com"
        """,
    ),
    (
        'importlib-metadata-smoke',
        r"""
from importlib_metadata import EntryPoint

entry = EntryPoint(name="demo", value="json:loads", group="console_scripts")
assert entry.module == "json"
assert entry.attr == "loads"
assert entry.load()("{\"answer\": 42}") == {"answer": 42}
        """,
    ),
    (
        'iniconfig-smoke',
        r"""
import tempfile
from pathlib import Path

from iniconfig import IniConfig

with tempfile.TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "demo.ini"
    path.write_text("[tool]\nname = staticpython\n", encoding="utf-8")
    config = IniConfig(str(path))
    assert config["tool"]["name"] == "staticpython"
    assert "tool" in config.sections
    assert list(config["tool"]) == ["name"]
        """,
    ),
    (
        'ipykernel-smoke',
        r"""
import importlib.util
import json
import sys
import tempfile
from pathlib import Path

import ipykernel
from ipykernel.kernelspec import write_kernel_spec
from jupyter_client import KernelManager
from jupyter_client.kernelspec import KernelSpecManager

with tempfile.TemporaryDirectory() as temp_dir:
    spec_dir = Path(temp_dir) / "kernels" / "python3"
    written_dir = Path(write_kernel_spec(spec_dir))
    assert written_dir == spec_dir

    kernel_json = json.loads((spec_dir / "kernel.json").read_text(encoding="utf-8"))
    has_debugpy = importlib.util.find_spec("debugpy") is not None
    assert kernel_json["argv"][0] == sys.executable
    assert "-m" in kernel_json["argv"]
    module_index = kernel_json["argv"].index("-m") + 1
    assert kernel_json["argv"][module_index] == "ipykernel_launcher"
    assert "-f" in kernel_json["argv"]
    assert kernel_json["metadata"]["debugger"] is has_debugpy
    assert (spec_dir / "logo-32x32.png").read_bytes().startswith(b"\x89PNG")
    assert (spec_dir / "logo-64x64.png").read_bytes().startswith(b"\x89PNG")
    assert "<svg" in (spec_dir / "logo-svg.svg").read_text(encoding="utf-8")

    prefix = Path(temp_dir) / "prefix"
    kernels_dir = prefix / "share" / "jupyter" / "kernels"
    kernel_spec_manager = KernelSpecManager(
        kernel_dirs=[str(kernels_dir)],
        ensure_native_kernel=False,
    )
    installed_path = Path(
        kernel_spec_manager.install_kernel_spec(
            str(spec_dir),
            kernel_name="python3",
            prefix=str(prefix),
        )
    )
    assert installed_path == kernels_dir / "python3"
    assert (installed_path / "kernel.json").exists()

    manager = KernelManager(
        kernel_name="python3",
        kernel_spec_manager=kernel_spec_manager,
    )
    manager.connection_file = str((Path(temp_dir) / "connection.json").resolve())
    manager.start_kernel()
    client = manager.client()
    client.start_channels()
    try:
        client.wait_for_ready(timeout=60)

        execute_id = client.execute("answer = 40 + 2\nanswer")
        while True:
            reply = client.get_shell_msg(timeout=60)
            if reply.get("parent_header", {}).get("msg_id") == execute_id:
                break
        assert reply["content"]["status"] == "ok"

        execute_result = None
        while True:
            message = client.get_iopub_msg(timeout=60)
            if message.get("parent_header", {}).get("msg_id") != execute_id:
                continue
            msg_type = message["header"]["msg_type"]
            if msg_type == "execute_result":
                execute_result = message["content"]["data"]["text/plain"]
            if msg_type == "status" and message["content"]["execution_state"] == "idle":
                break
        assert execute_result == "42"

        complete_id = client.complete("ans", 3)
        while True:
            completion = client.get_shell_msg(timeout=60)
            if completion.get("parent_header", {}).get("msg_id") == complete_id:
                break
        assert "answer" in completion["content"]["matches"]
    finally:
        client.stop_channels()
        manager.shutdown_kernel(now=True)
        """,
    ),
    (
        'ipython-smoke',
        r"""
import subprocess
import sys

from IPython.core.interactiveshell import InteractiveShell

shell = InteractiveShell.instance()
result = shell.run_cell("answer = 40 + 2", store_history=False)
assert result.success
assert shell.user_ns["answer"] == 42

completed = subprocess.run(
    [sys.executable, "-m", "IPython", "-c", "answer = 40 + 2; print(answer)"],
    capture_output=True,
    text=True,
    encoding="utf-8",
    errors="replace",
    timeout=120,
)
assert completed.returncode == 0, completed.stderr
assert completed.stdout.strip().endswith("42")
        """,
    ),
    (
        'ipython-pygments-lexers-smoke',
        r"""
from pygments import lex
from ipython_pygments_lexers import IPythonConsoleLexer

tokens = list(lex("In [1]: 1 + 1\n", IPythonConsoleLexer()))
assert tokens
assert any("In [1]" in value for _, value in tokens)
        """,
    ),
    (
        'isort-smoke',
        r"""
import isort
from isort.settings import Config

source = "import sys\nimport os\n"
sorted_source = isort.code(source, config=Config(profile="black"))
assert sorted_source.startswith("import os\nimport sys\n")
assert isort.check_code(sorted_source, config=Config(profile="black"))
        """,
    ),
    (
        'itsdangerous-smoke',
        r"""
from itsdangerous import BadSignature, URLSafeSerializer, URLSafeTimedSerializer

serializer = URLSafeSerializer("secret")
token = serializer.dumps({"value": 3})
assert serializer.loads(token) == {"value": 3}
try:
    serializer.loads(token + "x")
except BadSignature:
    pass
else:
    raise AssertionError("tampered token was accepted")
assert URLSafeTimedSerializer("secret").loads(URLSafeTimedSerializer("secret").dumps("ok")) == "ok"
        """,
    ),
    (
        'jedi-smoke',
        r"""
import jedi
from jedi.api.environment import InterpreterEnvironment

environment = InterpreterEnvironment()
script = jedi.Script("import math\nmath.sq", environment=environment)
completions = script.complete(2, 7)
names = {item.name for item in completions}
assert "sqrt" in names
inferred = jedi.Script("value = 42\nvalue", environment=environment).infer(2, 5)
assert inferred and inferred[0].name == "int"
local_completions = jedi.Interpreter("value.bit_", [{"value": 42}]).complete(1, 10)
assert {"bit_count", "bit_length"} <= {item.name for item in local_completions}
        """,
    ),
    (
        'jinja2-smoke',
        r"""
from jinja2 import DictLoader, Environment

env = Environment(loader=DictLoader({"demo.html": "Hello {{ name|upper }}"}))
template = env.get_template("demo.html")
assert template.render(name="codex") == "Hello CODEX"
assert env.from_string("{{ items|join(',') }}").render(items=[1, 2, 3]) == "1,2,3"
        """,
    ),
    (
        'jmespath-smoke',
        r"""
import jmespath

data = {"items": [{"name": "alpha", "value": 1}, {"name": "beta", "value": 2}]}
expression = jmespath.compile("items[?value>`1`].name | [0]")
assert expression.search(data) == "beta"
assert jmespath.search("length(items)", data) == 2
        """,
    ),
    (
        'joblib-smoke',
        r"""
from joblib import Parallel, delayed, hash
from joblib.memory import Memory

assert Parallel(n_jobs=1)(delayed(lambda value: value * value)(item) for item in range(4)) == [0, 1, 4, 9]
assert hash({"static": "python"})
memory = Memory(location=None, verbose=0)

@memory.cache
def add(a, b):
    return a + b

assert add(2, 5) == 7
        """,
    ),
    (
        'kiwisolver-smoke',
        r"""
import importlib.util

import kiwisolver as kiwi

assert importlib.util.find_spec("kiwisolver._cext").origin == "built-in"

x = kiwi.Variable("x")
y = kiwi.Variable("y")
solver = kiwi.Solver()
solver.addConstraint(x + y == 10)
solver.addConstraint(x - y == 2)
solver.updateVariables()
assert round(x.value(), 7) == 6.0
assert round(y.value(), 7) == 4.0

editable_x = kiwi.Variable("editable_x")
editable_y = kiwi.Variable("editable_y")
editable = kiwi.Solver()
editable.addConstraint(editable_x + editable_y == 10)
editable.addEditVariable(editable_x, kiwi.strength.strong)
editable.addEditVariable(editable_y, kiwi.strength.medium)
editable.suggestValue(editable_x, 8)
editable.suggestValue(editable_y, 1)
editable.updateVariables()
assert round(editable_x.value(), 7) == 8.0
assert round(editable_y.value(), 7) == 2.0
assert kiwi.__version__
        """,
    ),
    (
        'json5-smoke',
        r"""
import io

import json5

payload = json5.loads("{unquoted: 'value', trailing: [1,2,], flag: true}")
assert payload == {"unquoted": "value", "trailing": [1, 2], "flag": True}

buffer = io.StringIO("{answer: 42}\n")
assert json5.load(buffer) == {"answer": 42}

rendered = json5.dumps({"answer": 42}, quote_keys=True)
assert '"answer"' in rendered
assert json5.loads(rendered)["answer"] == 42
        """,
    ),
    (
        'jsonpickle-smoke',
        r"""
import jsonpickle


class Node:
    def __init__(self, name):
        self.name = name
        self.children = []


root = Node("root")
root.children.append(Node("leaf"))
encoded = jsonpickle.encode(root)
decoded = jsonpickle.decode(encoded)
assert decoded.name == "root"
assert decoded.children[0].name == "leaf"
assert "py/object" in encoded
        """,
    ),
    (
        'jsonschema-smoke',
        r"""
import jsonschema
from jsonschema import Draft7Validator, Draft202012Validator, FormatChecker, ValidationError

schema = {
    "type": "object",
    "properties": {
        "email": {"type": "string", "format": "email"},
        "items": {"type": "array", "items": {"type": "integer"}, "minItems": 2},
    },
    "required": ["email", "items"],
}

validator = Draft7Validator(schema, format_checker=FormatChecker())
validator.validate({"email": "codex@example.com", "items": [1, 2, 3]})
errors = sorted(validator.iter_errors({"email": "not-an-email", "items": [1]}), key=str)
assert len(errors) == 2

try:
    jsonschema.validate({"email": "codex@example.com", "items": [1, "bad"]}, schema)
except ValidationError as exc:
    assert exc.validator == "type"
else:
    raise AssertionError("jsonschema accepted invalid data")

schema_2020 = {
    "$schema": "https://json-schema.org/draft/2020-12/schema",
    "type": "object",
    "properties": {
        "name": {"type": "string"},
        "tags": {"type": "array", "items": {"type": "string"}, "minItems": 1},
    },
    "required": ["name"],
    "additionalProperties": False,
}
validator_2020 = Draft202012Validator(schema_2020)
validator_2020.validate({"name": "codex", "tags": ["a"]})
errors_2020 = sorted(
    validator_2020.iter_errors({"name": "codex", "tags": [1], "extra": 1}),
    key=lambda error: error.json_path,
)
assert len(errors_2020) == 2
assert {error.validator for error in errors_2020} == {"type", "additionalProperties"}
        """,
    ),
    (
        'jupyter-client-smoke',
        r"""
import json
import os
import sys
import tempfile
from pathlib import Path

from jupyter_client.connect import find_connection_file, write_connection_file
from jupyter_client.kernelspec import KernelSpec, KernelSpecManager
from jupyter_client.manager import KernelManager
from jupyter_client.provisioning.factory import KernelProvisionerFactory
from jupyter_client.session import Session

session = Session(key=b"staticpython-secret")
message = session.msg("execute_request", content={"code": "answer = 40 + 2"})
idents, frames = session.feed_identities(session.serialize(message))
assert idents == []
restored = session.deserialize(frames)
assert restored["header"]["msg_type"] == "execute_request"
assert restored["content"]["code"] == "answer = 40 + 2"

with tempfile.TemporaryDirectory() as temp_dir:
    connection_path = Path(temp_dir) / "kernel-staticpython.json"
    written_path, connection_info = write_connection_file(
        fname=str(connection_path),
        ip="127.0.0.1",
        key=b"secret",
        transport="tcp",
    )
    assert Path(written_path) == connection_path
    assert connection_info["ip"] == "127.0.0.1"
    assert connection_info["transport"] == "tcp"
    found_connection_file = Path(find_connection_file(connection_path.name, path=[temp_dir]))
    assert found_connection_file.name == connection_path.name
    assert found_connection_file.resolve() == connection_path.resolve()

    kernels_dir = Path(temp_dir) / "kernels"
    spec_dir = kernels_dir / "python3"
    spec_dir.mkdir(parents=True)
    spec = KernelSpec(
        argv=["python", "-m", "ipykernel_launcher", "-f", "{connection_file}"],
        display_name="StaticPython",
        language="python",
        resource_dir=str(spec_dir),
    )
    (spec_dir / "kernel.json").write_text(json.dumps(spec.to_dict(), indent=2), encoding="utf-8")

    kernel_spec_manager = KernelSpecManager(
        kernel_dirs=[str(kernels_dir)],
        ensure_native_kernel=False,
    )
    loaded_spec = kernel_spec_manager.get_kernel_spec("python3")
    assert loaded_spec.display_name == "StaticPython"
    factory = KernelProvisionerFactory()
    entry_point = factory._get_provisioner("local-provisioner")
    assert entry_point.name == "local-provisioner"
    assert entry_point.value == "jupyter_client.provisioning:LocalProvisioner"
    assert entry_point.group == KernelProvisionerFactory.GROUP_NAME

    manager = KernelManager(
        kernel_name="python3",
        kernel_spec_manager=kernel_spec_manager,
    )
    manager.connection_file = str(connection_path)
    formatted = manager.format_kernel_cmd(["--HistoryManager.enabled=False"])
    assert formatted[0] == sys.executable
    assert formatted[1:4] == ["-m", "ipykernel_launcher", "-f"]
    assert formatted[4] == str(connection_path.resolve())
    assert formatted[-1] == "--HistoryManager.enabled=False"
        """,
    ),
    (
        'jupyter-core-smoke',
        r"""
import os
import tempfile
from pathlib import Path

from jupyter_core.application import JupyterApp
from jupyter_core.paths import (
    ENV_JUPYTER_PATH,
    jupyter_config_dir,
    jupyter_data_dir,
    jupyter_path,
    jupyter_runtime_dir,
    secure_write,
)

assert isinstance(jupyter_config_dir(), str)
assert isinstance(jupyter_data_dir(), str)
assert isinstance(jupyter_runtime_dir(), str)
assert any(Path(item).name == "kernels" for item in jupyter_path("kernels"))
assert all(isinstance(item, str) for item in ENV_JUPYTER_PATH)

with tempfile.TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "connection.json"
    with secure_write(str(path)) as handle:
        handle.write("{}")
    assert path.read_text(encoding="utf-8") == "{}"

app = JupyterApp()
assert app.name == "jupyter"
assert isinstance(app.version, str)
assert os.path.isabs(jupyter_runtime_dir())
        """,
    ),
    (
        'jupyter-events-smoke',
        r"""
import io
import json
import logging

from jupyter_events import EventLogger, EventSchema
from jupyter_events.validators import EVENT_CORE_SCHEMA, EVENT_METASCHEMA, PROPERTY_METASCHEMA, validate_schema
from pythonjsonlogger import jsonlogger

assert EVENT_METASCHEMA["$id"].startswith("http://event.jupyter.org/")
assert EVENT_CORE_SCHEMA["$id"].startswith("http://event.jupyter.org/")
assert PROPERTY_METASCHEMA["$id"].startswith("http://event.jupyter.org/")

schema = {
    "$id": "https://events.example.com/staticpython/demo",
    "$schema": EVENT_METASCHEMA["$id"],
    "version": 1,
    "title": "StaticPython demo event",
    "description": "Smoke test schema.",
    "type": "object",
    "properties": {
        "status": {"type": "string"},
        "value": {"type": "integer"},
    },
    "required": ["status"],
    "additionalProperties": False,
}

validate_schema(schema)
event_schema = EventSchema(schema)
assert event_schema.id == schema["$id"]

stream = io.StringIO()
handler = logging.StreamHandler(stream)
logger = EventLogger()
logger.register_handler(handler)
assert isinstance(handler.formatter, jsonlogger.JsonFormatter)
logger.register_event_schema(schema)
assert schema["$id"] in logger.schemas.schema_ids

capsule = logger.emit(schema_id=schema["$id"], data={"status": "ok", "value": 42})
handler.flush()
payload = json.loads(stream.getvalue())

assert capsule["__schema__"] == schema["$id"]
assert payload["status"] == "ok"
assert payload["value"] == 42
assert payload["__schema__"] == schema["$id"]
assert payload["__metadata_version__"] == 1
        """,
    ),
    (
        'jupyter-lsp-smoke',
        r"""
from jupyter_lsp.manager import LanguageServerManager
from jupyter_lsp.schema import LANGUAGE_SERVER_SPEC, LANGUAGE_SERVER_SPEC_MAP, SERVERS_RESPONSE

spec = {
    "version": 2,
    "argv": ["pylsp"],
    "languages": ["python"],
    "display_name": "Python LSP",
    "mime_types": ["text/x-python"],
    "requires_documents_on_disk": False,
}

LANGUAGE_SERVER_SPEC.validate(spec)
LANGUAGE_SERVER_SPEC_MAP.validate({"pylsp": spec})
SERVERS_RESPONSE.validate({"version": 2, "sessions": {}, "specs": {"pylsp": spec}})

manager = LanguageServerManager()
assert manager.virtual_documents_dir == ".virtual_documents"
assert manager.language_servers == {}
assert manager.conf_d_language_servers == {}
        """,
    ),
    (
        'jupyter-server-smoke',
        r"""
from pathlib import Path
from types import SimpleNamespace

import jupyter_server
from jupyter_server.base.handlers import AuthenticatedHandler
from jupyter_server.extension.application import ExtensionApp
from jupyter_server.serverapp import ServerApp
from jupyter_server.utils import url_path_join

template_roots = [Path(path) for path in jupyter_server.DEFAULT_TEMPLATE_PATH_LIST]

assert Path(jupyter_server.DEFAULT_STATIC_FILES_PATH).name == "static"
assert Path(jupyter_server.DEFAULT_EVENTS_SCHEMA_PATH).name == "event_schemas"
assert any(path.name == "templates" for path in template_roots)
assert url_path_join("/base/", "api", "status") == "/base/api/status"

app = ServerApp()
assert app.default_url == "/"
assert app.contents_manager_class is not None
app.gateway_config = SimpleNamespace(gateway_enabled=False)
assert app.kernel_manager_class is not None
assert app.session_manager_class is not None
assert app.kernel_spec_manager_class is not None
assert app.kernel_websocket_connection_class is not None
assert any(Path(path).name == "static" for path in app.static_file_path)
assert issubclass(ExtensionApp, object)
assert callable(AuthenticatedHandler.set_default_headers)
        """,
    ),
    (
        'jupyterlab-smoke',
        r"""
import json
import os
import socket
import subprocess
import sys
import tempfile
import time
import urllib.request
from pathlib import Path

from jupyterlab.commands import get_app_dir, get_user_settings_dir, get_workspaces_dir
from jupyterlab.coreconfig import CoreConfig
from jupyterlab.extensions import MANAGERS
from jupyterlab.labapp import LabApp
from jupyterlab_server.settings_utils import get_settings

assert "readonly" in MANAGERS
assert "pypi" in MANAGERS

app_dir = Path(get_app_dir())
assert app_dir.name == "lab"
assert "user-settings" in get_user_settings_dir()
assert "workspaces" in get_workspaces_dir()

config = CoreConfig()
assert config.static_dir == "../static"
assert "@jupyterlab/application-extension" in config.extensions

app = LabApp()
assert app.default_url == "/lab"
assert Path(app.static_dir).name == "static"
assert Path(app.templates_dir).name == "static"
assert Path(app.schemas_dir).name == "schemas"
settings = get_settings(
    app.app_settings_dir,
    app.schemas_dir,
    app.user_settings_dir,
    "@jupyterlab/apputils-extension:themes",
    overrides={},
    labextensions_path=app.extra_labextensions_path + app.labextensions_path,
)
settings, warnings = settings
assert warnings == [None] or warnings == []
assert settings["id"] == "@jupyterlab/apputils-extension:themes"
assert "theme" in settings["schema"]["title"].lower()
all_settings, all_warnings = get_settings(
    app.app_settings_dir,
    app.schemas_dir,
    app.user_settings_dir,
    overrides={},
    labextensions_path=app.extra_labextensions_path + app.labextensions_path,
    ids_only=True,
)
assert "@jupyterlab/apputils-extension:themes" in {entry["id"] for entry in all_settings["settings"]}


def reserve_port() -> int:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.bind(("127.0.0.1", 0))
    try:
        return sock.getsockname()[1]
    finally:
        sock.close()


with tempfile.TemporaryDirectory() as temp_dir:
    temp_root = Path(temp_dir)
    work_dir = temp_root / "work"
    work_dir.mkdir()
    port = reserve_port()
    token = "staticpython-jupyterlab"
    env = os.environ.copy()
    env.update(
        {
            "JUPYTER_CONFIG_DIR": str(temp_root / "config"),
            "JUPYTER_DATA_DIR": str(temp_root / "data"),
            "JUPYTER_RUNTIME_DIR": str(temp_root / "runtime"),
            "JUPYTERLAB_SETTINGS_DIR": str(temp_root / "lab" / "user-settings"),
            "JUPYTERLAB_WORKSPACES_DIR": str(temp_root / "lab" / "workspaces"),
            "PYTHONUTF8": "1",
            "PYTHONIOENCODING": "utf-8",
        }
    )
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    command = [
        sys.executable,
        "-m",
        "jupyterlab",
        "--ServerApp.ip=127.0.0.1",
        f"--ServerApp.port={port}",
        "--ServerApp.port_retries=0",
        "--ServerApp.open_browser=False",
        f"--ServerApp.root_dir={work_dir}",
        f"--ServerApp.token={token}",
        "--ServerApp.password=",
    ]
    process = subprocess.Popen(
        command,
        cwd=str(work_dir),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=creationflags,
    )
    output = ""
    try:
        url = f"http://127.0.0.1:{port}/lab?token={token}"
        response_text = None
        last_error = None
        deadline = time.time() + 90
        while time.time() < deadline:
            if process.poll() is not None:
                break
            try:
                with urllib.request.urlopen(url, timeout=5) as response:
                    response_text = response.read().decode("utf-8", errors="replace")
                    status_code = response.status
                break
            except Exception as exc:
                last_error = exc
                time.sleep(0.5)
        if response_text is None:
            if process.poll() is None:
                process.terminate()
                try:
                    process.wait(timeout=15)
                except subprocess.TimeoutExpired:
                    process.kill()
                    process.wait(timeout=15)
            if process.stdout is not None:
                output = process.stdout.read()
            raise AssertionError(
                f"failed to load /lab within timeout: {last_error!r}\n"
                f"process return code: {process.returncode}\n"
                f"jupyter output:\n{output[-8000:]}"
            )
        assert status_code == 200
        assert "<title>JupyterLab</title>" in response_text
        assert "jupyter-config-data" in response_text
    finally:
        if process.poll() is None:
            process.terminate()
            try:
                process.wait(timeout=15)
            except subprocess.TimeoutExpired:
                process.kill()
                process.wait(timeout=15)
        if process.stdout is not None:
            output = process.stdout.read()
        if process.returncode not in (0, -15, 1):
            raise AssertionError(
                f"jupyterlab server exited unexpectedly with code {process.returncode}\n"
                f"jupyter output:\n{output[-8000:]}"
            )
        """,
    ),
    (
        'jupyterlab-pygments-smoke',
        r"""
from pygments import highlight
from pygments.formatters import HtmlFormatter
from pygments.lexers import PythonLexer

from jupyterlab_pygments import JupyterStyle

html = highlight("print('staticpython')\n", PythonLexer(), HtmlFormatter(style=JupyterStyle))
assert "jp-RenderedHTMLCommon" not in html or isinstance(html, str)
assert "staticpython" in html
assert "highlight" in html
        """,
    ),
    (
        'jupyterlab-server-smoke',
        r"""
from pathlib import Path

from jupyterlab_server import LabServerApp
from jupyterlab_server.config import get_page_config
from jupyterlab_server.settings_utils import _get_user_settings
from jupyterlab_server.workspaces_handler import slugify
import jupyterlab_server

app = LabServerApp()
assert app.default_url == "/lab"
assert app.settings_url == "/lab/api/settings/"
assert app.translations_api_url == "/lab/api/translations/"
assert app.workspaces_api_url == "/lab/api/workspaces/"
assert app.themes_url == "/lab/api/themes/"
assert app.licenses_url == "/lab/api/licenses/"
assert app.templates_dir == ""
assert app.schemas_dir == ""

settings = _get_user_settings(str(Path.cwd()), "@jupyterlab/apputils-extension:themes", {"type": "object"})
assert settings["raw"] == "{}" or settings["raw"] == {}
assert settings["settings"] == {}

page_config = get_page_config([], logger=app.log)
assert isinstance(page_config, dict)
assert isinstance(page_config.get("federated_extensions"), list)
if page_config["federated_extensions"]:
    assert any(entry["name"] == "@jupyter-notebook/lab-extension" for entry in page_config["federated_extensions"])
assert slugify("/StaticPython Workspace") != ""
        """,
    ),
    (
        'lark-parser-smoke',
        r"""
from lark import Lark, Transformer, v_args

parser = Lark("start: WORD NUMBER\n%import common.WORD\n%import common.NUMBER\n%ignore \" \"", parser="lalr")
tree = parser.parse("staticpython 42")
assert tree.data == "start"
assert [child.value for child in tree.children] == ["staticpython", "42"]

@v_args(inline=True)
class Calc(Transformer):
    def number(self, token):
        return int(token)
    def add(self, left, right):
        return left + right

calc = Lark("start: sum\n?sum: number -> number | sum \"+\" number -> add\nnumber: NUMBER\n%import common.NUMBER\n%ignore \" \"", parser="lalr")
result = Calc().transform(calc.parse("2 + 40"))
assert result.children == [42]
        """,
    ),
    (
        'loguru-smoke',
        r"""
import io
from loguru import logger

buffer = io.StringIO()
handler_id = logger.add(buffer, format="{level}:{message}")
try:
    logger.warning("demo")
    logger.bind(component="verify").info("bound")
finally:
    logger.remove(handler_id)
output = buffer.getvalue()
assert "WARNING:demo" in output
assert "INFO:bound" in output
        """,
    ),
    (
        'mako-smoke',
        r"""
from mako.lookup import TemplateLookup
from mako.template import Template

assert Template("hello ${name}").render(name="codex") == "hello codex"
lookup = TemplateLookup()
lookup.put_string("demo", "${x + y}")
assert lookup.get_template("demo").render(x=2, y=3) == "5"
        """,
    ),
    (
        'markdown-smoke',
        r"""
import markdown

html = markdown.markdown("# Title\n\n- a\n- b", extensions=["extra"])
assert "<h1>Title</h1>" in html
assert "<li>a</li>" in html
        """,
    ),
    (
        'markdown-it-smoke',
        r"""
from markdown_it import MarkdownIt

md = MarkdownIt()
html = md.render("**bold**\n\n[link](https://example.com)")
assert "<strong>bold</strong>" in html
assert 'href="https://example.com"' in html
        """,
    ),
    (
        'markupsafe-smoke',
        r"""
import importlib.util
import markupsafe
import markupsafe._speedups as speedups

assert importlib.util.find_spec("markupsafe._speedups").origin == "built-in"
assert speedups._escape_inner("<codex & static>") == "&lt;codex &amp; static&gt;"
assert str(markupsafe.escape("<hello>")) == "&lt;hello&gt;"
assert str(markupsafe.Markup("<b>{}</b>").format("x&y")) == "<b>x&amp;y</b>"
        """,
    ),
    (
        'marshmallow-smoke',
        r"""
from marshmallow import Schema, ValidationError, fields

class UserSchema(Schema):
    name = fields.Str(required=True)
    age = fields.Int(required=True)

schema = UserSchema()
loaded = schema.load({"name": "Ada", "age": "42"})
assert loaded == {"name": "Ada", "age": 42}
assert schema.dump(loaded) == {"name": "Ada", "age": 42}
try:
    schema.load({"age": "bad"})
except ValidationError as exc:
    assert "name" in exc.messages and "age" in exc.messages
else:
    raise AssertionError("marshmallow accepted invalid input")
        """,
    ),
    (
        'matplotlib-inline-smoke',
        r"""
from matplotlib_inline.config import InlineBackend

backend = InlineBackend.instance()
assert backend.figure_formats == {"png"}
backend.figure_format = "svg"
assert backend.figure_formats == {"svg"}
assert backend.print_figure_kwargs["bbox_inches"] == "tight"
        """,
    ),
    (
        'matplotlib-smoke',
        r"""
import importlib.util
import io
import os
import tempfile

_staticpython_mpl_tmpdir = tempfile.TemporaryDirectory(prefix="staticpython-mpl-")
os.environ.setdefault("MPLCONFIGDIR", _staticpython_mpl_tmpdir.name)

import matplotlib

matplotlib.use("Agg", force=True)

for name in (
    "matplotlib.ft2font",
    "matplotlib._path",
    "matplotlib._image",
    "matplotlib._qhull",
    "matplotlib._tri",
    "matplotlib._c_internal_utils",
    "matplotlib.backends._backend_agg",
):
    spec = importlib.util.find_spec(name)
    assert spec is not None, name
    assert spec.origin == "built-in", (name, spec.origin)

from matplotlib import ft2font
from matplotlib import pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

assert ft2font.__freetype_version__
assert ft2font.__freetype_build_type__ == "local"
assert Axes3D.__name__ == "Axes3D"

data_path = matplotlib.get_data_path()
assert data_path.endswith("mpl-data"), data_path

fig, ax = plt.subplots(figsize=(2, 1.5), dpi=80)
ax.plot([0, 1, 2], [0, 1, 0], marker="o")
ax.set_title("StaticPython")
ax.fill_between([0, 1, 2], [0, 0.25, 0], alpha=0.2)

buffer = io.BytesIO()
fig.savefig(buffer, format="png")
plt.close(fig)

payload = buffer.getvalue()
assert payload.startswith(b"\x89PNG\r\n\x1a\n")
assert len(payload) > 1024
        """,
    ),
    (
        'mdurl-smoke',
        r"""
from mdurl import decode, encode, parse

assert encode("a b") == "a%20b"
assert decode("a%20b") == "a b"
parsed = parse("https://user:pass@example.com:443/a?b=1#frag")
assert parsed.protocol == "https:"
assert parsed.hostname == "example.com"
assert parsed.pathname == "/a"
        """,
    ),
    (
        'mistune-smoke',
        r"""
import mistune

html = mistune.html("# Title\n\n**bold**")
assert "<h1>Title</h1>" in html
assert "<strong>bold</strong>" in html
renderer = mistune.create_markdown()
assert "<p>demo</p>" in renderer("demo")
        """,
    ),
    (
        'more-itertools-smoke',
        r"""
from more_itertools import chunked, flatten, pairwise

assert list(chunked([1, 2, 3, 4], 2)) == [[1, 2], [3, 4]]
assert list(flatten([[1, 2], [3]])) == [1, 2, 3]
assert list(pairwise([1, 2, 3])) == [(1, 2), (2, 3)]
        """,
    ),
    (
        'mpmath-smoke',
        r"""
import mpmath as mp

mp.mp.dps = 30
assert mp.sqrt(81) == 9
assert str(mp.factorial(5)) == "120.0"
assert abs(mp.sin(mp.pi / 2) - 1) < mp.mpf("1e-25")
        """,
    ),
    (
        'msgpack-smoke',
        r"""
import importlib.util
import msgpack
import msgpack._cmsgpack as cmsgpack

assert importlib.util.find_spec("msgpack._cmsgpack").origin == "built-in"
payload = {"name": "codex", "items": [1, 2, 3], "binary": b"data"}
packed = msgpack.packb(payload, use_bin_type=True)
assert msgpack.unpackb(packed, raw=False) == payload
packer = cmsgpack.Packer(use_bin_type=True)
assert msgpack.unpackb(packer.pack(["static", 13]), raw=False) == ["static", 13]
        """,
    ),
    (
        'multidict-smoke',
        r"""
import importlib.util
import multidict
import multidict._compat
import multidict._multidict

assert importlib.util.find_spec("multidict._multidict").origin == "built-in"
assert multidict._compat.USE_EXTENSIONS is True
md = multidict.MultiDict([("x", "1"), ("x", "2")])
assert md.getall("x") == ["1", "2"]
md.add("y", "3")
assert list(md.items()) == [("x", "1"), ("x", "2"), ("y", "3")]
ci = multidict.CIMultiDict({"Content-Type": "text/plain"})
assert ci["content-type"] == "text/plain"
assert multidict.getversion(md) >= 0
        """,
    ),
    (
        'mypy-extensions-smoke',
        r"""
from mypy_extensions import (
    Arg,
    DefaultArg,
    DefaultNamedArg,
    FlexibleAlias,
    KwArg,
    NamedArg,
    TypedDict,
    VarArg,
    i16,
    i32,
    i64,
    mypyc_attr,
    trait,
    u8,
)

class Demo(TypedDict):
    value: int

FunctionalDemo = TypedDict("FunctionalDemo", {"name": str, "count": int}, total=False)

assert Demo(value=42)["value"] == 42
assert FunctionalDemo(name="codex") == {"name": "codex"}
assert "value" in Demo.__annotations__
assert FunctionalDemo.__total__ is False
assert Arg(int) is int
assert DefaultArg(str) is str
assert NamedArg(float) is float
assert DefaultNamedArg(bytes) is bytes
assert VarArg(tuple) is tuple
assert KwArg(dict) is dict
assert FlexibleAlias[int, str][float, bytes] is str
assert i64("42") == 42 and i32(7) == 7 and i16(8) == 8 and u8(9) == 9
assert isinstance(1, i64) and isinstance(1, i32) and isinstance(1, i16) and isinstance(1, u8)
assert trait(lambda: "ok")() == "ok"
assert mypyc_attr("allow_interpreted_subclasses")(lambda: "ok")() == "ok"
        """,
    ),
    (
        'narwhals-smoke',
        r"""
import narwhals
from narwhals.dependencies import is_into_dataframe

assert callable(narwhals.from_native)
assert narwhals.Int64.__name__ == "Int64"
assert is_into_dataframe({"a": [1, 2]}) is False
        """,
    ),
    (
        'nbclient-smoke',
        r"""
from nbclient import NotebookClient
from nbformat import v4

notebook = v4.new_notebook()
notebook.cells.append(v4.new_markdown_cell("# StaticPython"))
notebook.cells.append(v4.new_code_cell("answer = 40 + 2\nanswer"))

client = NotebookClient(
    notebook,
    kernel_name="python3",
    timeout=60,
)
executed = client.execute()
assert executed.cells[0].source == "# StaticPython"
assert executed.cells[1].execution_count == 1
assert executed.cells[1].outputs
output = executed.cells[1].outputs[0]
assert output["output_type"] == "execute_result"
assert output["data"]["text/plain"] == "42"
        """,
    ),
    (
        'nbconvert-smoke',
        r"""
from nbformat import v4
from nbconvert.exporters import HTMLExporter, ScriptExporter
from nbconvert.exporters.base import get_export_names, get_exporter

notebook = v4.new_notebook()
notebook.cells.append(v4.new_markdown_cell("# StaticPython"))
notebook.cells.append(v4.new_code_cell("answer = 40 + 2\nanswer"))

export_names = set(get_export_names())
assert {"notebook"} <= export_names
assert get_exporter("html").__name__ == "HTMLExporter"
script_exporter_class = get_exporter("script")
assert script_exporter_class.__name__ in {"ScriptExporter", "PythonExporter"}

html_exporter = HTMLExporter()
html_body, html_resources = html_exporter.from_notebook_node(notebook)
assert "<h1" in html_body and "StaticPython" in html_body
assert html_exporter.get_template_names()[0] == "lab"
assert '<div class="highlight hl-ipython3"><pre>' in html_body
assert "answer" in html_body and "40" in html_body and "2" in html_body
assert isinstance(html_resources, dict)

script_exporter = ScriptExporter()
script_body, script_resources = script_exporter.from_notebook_node(notebook)
assert "answer = 40 + 2" in script_body
assert script_resources.get("output_extension") in {".txt", ".py"}
        """,
    ),
    (
        'nbformat-smoke',
        r"""
import io

import nbformat
from nbformat import v4

nb = v4.new_notebook()
nb.cells.append(v4.new_markdown_cell("# StaticPython"))
nb.cells.append(v4.new_code_cell("answer = 40 + 2"))

nbformat.validate(nb)
encoded = nbformat.writes(nb)
decoded = nbformat.reads(encoded, as_version=4)
assert decoded.cells[0].source == "# StaticPython"
assert decoded.cells[1].cell_type == "code"

buffer = io.StringIO()
nbformat.write(decoded, buffer)
buffer.seek(0)
roundtripped = nbformat.read(buffer, as_version=4)
assert roundtripped.nbformat == 4
assert nbformat.validator.isvalid(roundtripped)
        """,
    ),
    (
        'networkx-smoke',
        r"""
import networkx as nx

graph = nx.Graph()
graph.add_weighted_edges_from([("a", "b", 2), ("b", "c", 3), ("a", "c", 10)])
assert nx.shortest_path(graph, "a", "c", weight="weight") == ["a", "b", "c"]
assert nx.shortest_path_length(graph, "a", "c", weight="weight") == 5
assert nx.is_connected(graph)
        """,
    ),
    (
        'notebook-smoke',
        r"""
import json
from pathlib import Path

from notebook import _jupyter_labextension_paths
from notebook.app import JupyterNotebookApp
from jupyterlab_server.config import get_page_config
from jupyterlab_server.settings_utils import get_settings

paths = _jupyter_labextension_paths()
assert paths == [{"src": "labextension", "dest": "@jupyter-notebook/lab-extension"}]

app = JupyterNotebookApp()
assert app.default_url == "/tree"
assert Path(app.static_dir).name == "static"
assert Path(app.templates_dir).name == "templates"
assert Path(app.schemas_dir).name == "schemas"
assert Path(app.app_dir).name == "lab"
page_config = get_page_config(app.extra_labextensions_path + app.labextensions_path, logger=app.log)
extension_names = {entry["name"] for entry in page_config["federated_extensions"]}
assert "@jupyter-notebook/lab-extension" in extension_names
entry = next(entry for entry in page_config["federated_extensions"] if entry["name"] == "@jupyter-notebook/lab-extension")
assert page_config["disabledExtensions"] == []
assert entry["load"].startswith("static/")
assert entry["style"] == "./style"
settings = get_settings(
    app.app_settings_dir,
    app.schemas_dir,
    app.user_settings_dir,
    "@jupyter-notebook/lab-extension:interface-switcher",
    overrides={},
    labextensions_path=app.extra_labextensions_path + app.labextensions_path,
)
settings, warnings = settings
assert warnings == [None] or warnings == []
assert settings["id"] == "@jupyter-notebook/lab-extension:interface-switcher"
assert settings["schema"]["title"]
        """,
    ),
    (
        'notebook-shim-smoke',
        r"""
from traitlets.config import Config

from jupyter_server.extension.application import ExtensionApp
from notebook_shim.shim import NotebookConfigShimMixin


class DemoApp(NotebookConfigShimMixin, ExtensionApp):
    pass


cfg = Config(
    {
        "NotebookApp": {"allow_remote_access": True},
        "ServerApp": {"port": 9999},
        "DemoApp": {"default_url": "/demo"},
    }
)

app = DemoApp()
shimmed = app.shim_config_from_notebook_to_jupyter_server(cfg)

assert shimmed["ServerApp"]["allow_remote_access"] is True
assert shimmed["ServerApp"]["port"] == 9999
assert shimmed["DemoApp"]["default_url"] == "/demo"
        """,
    ),
    (
        'numpy-smoke',
        r"""
import importlib.util
import io
import pickle

import numpy as np

assert importlib.util.find_spec("numpy._core._multiarray_umath").origin == "built-in"
assert np.__version__.startswith("2.4.")
assert np.dtype(np.float64).itemsize == 8

matrix = np.arange(6, dtype=np.int64).reshape(2, 3)
vector = np.array([10, 20, 30], dtype=np.int64)
assert matrix.shape == (2, 3)
assert (matrix + vector).tolist() == [[10, 21, 32], [13, 24, 35]]
assert np.dot(np.array([1, 2, 3]), np.array([4, 5, 6])) == 32
assert np.mean(matrix) == 2.5
assert np.array_equal(np.frombuffer(b"\x01\x00\x02\x00", dtype=np.uint16), np.array([1, 2], dtype=np.uint16))

buffer = io.BytesIO()
np.save(buffer, matrix)
buffer.seek(0)
loaded = np.load(buffer)
assert np.array_equal(loaded, matrix)

solution = np.linalg.solve(
    np.array([[2.0, 0.0], [0.0, 4.0]], dtype=np.float64),
    np.array([4.0, 8.0], dtype=np.float64),
)
assert np.allclose(solution, np.array([2.0, 2.0], dtype=np.float64))

fft = np.fft.rfft(np.array([0.0, 1.0, 0.0, -1.0], dtype=np.float64))
assert np.allclose(fft, np.array([0.0 + 0.0j, 0.0 - 2.0j, 0.0 + 0.0j]))

rng_a = np.random.default_rng(12345)
rng_b = np.random.default_rng(12345)
sample_a = rng_a.integers(0, 100, size=6)
sample_b = rng_b.integers(0, 100, size=6)
assert sample_a.shape == (6,)
assert sample_a.tolist() == sample_b.tolist()
assert all(0 <= value < 100 for value in sample_a.tolist())

restored = pickle.loads(pickle.dumps(matrix))
assert np.array_equal(restored, matrix)
assert np.__config__ is not None
        """,
    ),
    (
        'oauthlib-smoke',
        r"""
from oauthlib.oauth2 import WebApplicationClient

client = WebApplicationClient("client-id")
url = client.prepare_request_uri(
    "https://example.com/authorize",
    redirect_uri="https://client.example/callback",
    scope=["profile", "email"],
    state="state",
)
assert "client_id=client-id" in url
assert "response_type=code" in url
assert "scope=profile+email" in url
        """,
    ),
    (
        'openpyxl-smoke',
        r"""
import io
import openpyxl
from openpyxl.styles import Font

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Data"
sheet["A1"] = "name"
sheet["B1"] = "value"
sheet["A2"] = "codex"
sheet["B2"] = "=SUM(20,22)"
sheet["A1"].font = Font(bold=True)
sheet.merge_cells("C1:D1")
sheet["C1"] = "merged"
buffer = io.BytesIO()
workbook.save(buffer)
buffer.seek(0)
loaded = openpyxl.load_workbook(buffer, data_only=False)
assert loaded["Data"]["A1"].font.bold is True
assert loaded["Data"]["A2"].value == "codex"
assert loaded["Data"]["B2"].value == "=SUM(20,22)"
assert "C1:D1" in [str(item) for item in loaded["Data"].merged_cells.ranges]
        """,
    ),
    (
        'outcome-smoke',
        r"""
import outcome

value = outcome.capture(lambda: 42)
assert isinstance(value, outcome.Value)
assert value.unwrap() == 42

error = outcome.capture(lambda: 1 / 0)
assert isinstance(error, outcome.Error)
try:
    error.unwrap()
except ZeroDivisionError:
    pass
else:
    raise AssertionError("outcome did not re-raise captured error")
        """,
    ),
    (
        'overrides-smoke',
        r"""
from overrides import override


class Base:
    def value(self):
        return "base"


class Child(Base):
    @override
    def value(self):
        return "child"


assert Child().value() == "child"
assert Base().value() == "base"
        """,
    ),
    (
        'packaging-smoke',
        r"""
from packaging.markers import Marker
from packaging.requirements import Requirement
from packaging.specifiers import SpecifierSet
from packaging.version import Version

assert Version("1.2.3") < Version("2.0")
assert Version("1.5") in SpecifierSet(">=1,<2")
req = Requirement("demo[extra]>=1; python_version >= '3.8'")
assert req.name == "demo" and "extra" in req.extras
assert Marker("python_version >= '3.8'").evaluate()
        """,
    ),
    (
        'pandas-smoke',
        r"""
import importlib.util
import io

import numpy as np
import pandas as pd

assert importlib.util.find_spec("pandas._libs.algos").origin == "built-in"
assert importlib.util.find_spec("pandas._libs._cyutility").origin == "built-in"
assert importlib.util.find_spec("pandas._libs.tslibs.base").origin == "built-in"
assert importlib.util.find_spec("pandas._libs.window.aggregations").origin == "built-in"
assert pd.__version__.startswith("3.0.")

left = pd.DataFrame(
    {
        "id": [1, 2, 3],
        "value": [10, 20, 30],
        "when": pd.to_datetime(
            ["2024-01-01T00:00:00Z", "2024-01-01T01:00:00Z", "2024-01-02T00:00:00Z"],
            utc=True,
        ),
    }
)
right = pd.DataFrame({"id": [1, 2, 4], "name": ["a", "b", "d"]})
merged = left.merge(right, on="id", how="left")
assert merged["name"].iloc[0] == "a"
assert merged["name"].iloc[1] == "b"
assert pd.isna(merged["name"].iloc[2])

grouped = merged.groupby(merged["when"].dt.day)["value"].sum()
assert grouped.to_dict() == {1: 30, 2: 30}

csv_buffer = io.StringIO()
merged.to_csv(csv_buffer, index=False)
csv_buffer.seek(0)
roundtrip = pd.read_csv(csv_buffer)
assert roundtrip.shape == (3, 4)
assert roundtrip["value"].sum() == 60

json_frame = pd.read_json(io.StringIO('[{"id": 1, "value": 2}, {"id": 2, "value": 3}]'))
assert json_frame["value"].sum() == 5

pivot = pd.DataFrame(
    {
        "kind": ["a", "a", "b"],
        "column": ["x", "y", "x"],
        "value": [1, 2, 3],
    }
).pivot_table(index="kind", columns="column", values="value", aggfunc="sum")
assert pivot.loc["a", "x"] == 1
assert pivot.loc["a", "y"] == 2
assert pivot.loc["b", "x"] == 3

rolling = pd.Series([1, 2, 3, 4], dtype="float64").rolling(2).sum()
assert pd.isna(rolling.iloc[0])
assert rolling.iloc[1:].tolist() == [3.0, 5.0, 7.0]

encoded = left.to_json()
decoded = pd.read_json(io.StringIO(encoded))
assert decoded.shape[0] == 3
assert sorted(decoded.columns.tolist()) == ["id", "value", "when"]

timestamp = pd.Timestamp("2024-01-02T03:04:05", tz="UTC")
assert timestamp.tz is not None
date_range = pd.date_range("2024-01-01", periods=3, tz="UTC")
assert len(date_range) == 3

series = pd.Series([1.0, None, 3.0]).fillna(2.0)
assert series.tolist() == [1.0, 2.0, 3.0]
assert np.array_equal(series.to_numpy(), np.array([1.0, 2.0, 3.0]))
        """,
    ),
    (
        'pandocfilters-smoke',
        r"""
import json

from pandocfilters import Para, Str, applyJSONFilters, stringify

document = {
    "pandoc-api-version": [1, 22],
    "meta": {},
    "blocks": [Para([Str("StaticPython")])],
}

def uppercase(key, value, fmt, meta):
    if key == "Str":
        return Str(value.upper())
    return None

result = json.loads(applyJSONFilters([uppercase], json.dumps(document), format="html"))
assert result["blocks"][0]["c"][0]["c"] == "STATICPYTHON"
assert stringify(result) == "STATICPYTHON"
        """,
    ),
    (
        'parso-smoke',
        r"""
import parso

module = parso.parse("def func(value):\n    return value + 1\n")
function = module.children[0]
assert function.name.value == "func"
assert function.get_code().startswith("def func")
assert "return value + 1" in function.get_code()
        """,
    ),
    (
        'pathspec-smoke',
        r"""
from pathspec import PathSpec

spec = PathSpec.from_lines("gitwildmatch", ["*.pyc", "build/"])
assert spec.match_file("demo.pyc")
assert spec.match_file("build/output.txt")
assert not spec.match_file("src/demo.py")
        """,
    ),
    (
        'pillow-smoke',
        r"""
import importlib.util
from io import BytesIO

from PIL import Image, ImageChops, ImageDraw, ImageFilter, ImageMath, ImageMorph, ImageOps, ImageStat

assert importlib.util.find_spec("PIL._imaging").origin == "built-in"
assert importlib.util.find_spec("PIL._imagingmath").origin == "built-in"
assert importlib.util.find_spec("PIL._imagingmorph").origin == "built-in"

Image.preinit()
Image.init()
registered = Image.registered_extensions()
assert registered[".bmp"] == "BMP"
assert registered[".png"] == "PNG"

image = Image.new("RGB", (6, 6), "navy")
draw = ImageDraw.Draw(image)
draw.rectangle((1, 1, 4, 4), fill=(255, 0, 0))

assert image.getpixel((1, 1)) == (255, 0, 0)
assert image.crop((1, 1, 5, 5)).size == (4, 4)
assert image.resize((2, 2)).size == (2, 2)
assert image.convert("L").mode == "L"
assert ImageOps.mirror(image).size == image.size
assert ImageChops.difference(image, image).getbbox() is None
assert image.filter(ImageFilter.BLUR).size == image.size
assert ImageStat.Stat(image).sum[0] > 0

thumbnail = image.copy()
thumbnail.thumbnail((3, 3))
assert thumbnail.size == (3, 3)

alpha = Image.new("RGBA", (2, 2), (10, 20, 30, 128))
composited = Image.alpha_composite(Image.new("RGBA", (2, 2), (0, 0, 0, 255)), alpha)
assert composited.mode == "RGBA"

math_result = ImageMath.lambda_eval(lambda args: args["a"] + 1, a=Image.new("L", (1, 1), 41))
assert math_result.getpixel((0, 0)) == 42

morph = ImageMorph.MorphOp(op_name="dilation4")
assert morph.get_on_pixels(Image.new("L", (3, 3), 0)) == []

for image_format in ("BMP", "PNG"):
    buffer = BytesIO()
    image.save(buffer, format=image_format)
    buffer.seek(0)
    loaded = Image.open(buffer)
    loaded.load()
    assert loaded.size == (6, 6)
    assert loaded.mode == "RGB"
        """,
    ),
    (
        'platformdirs-smoke',
        r"""
from platformdirs import user_cache_dir, user_config_dir

cache_dir = user_cache_dir("StaticPython", "StaticPython")
config_dir = user_config_dir("StaticPython", "StaticPython")
assert "StaticPython" in cache_dir
assert "StaticPython" in config_dir
        """,
    ),
    (
        'plotly-smoke',
        r"""
import json
import plotly.graph_objects as go
import plotly.data as plotly_data
import plotly.io as pio
from plotly.offline import get_plotlyjs

figure = go.Figure(data=[go.Scatter(x=[1, 2], y=[3, 4], mode="lines+markers")])
payload = json.loads(figure.to_json())
assert payload["data"][0]["type"] == "scatter"
assert "plotly" in pio.templates
figure.update_layout(template="plotly")
assert figure.layout.template is not None
assert "Plotly" in get_plotlyjs()[:2000]
assert len(plotly_data.election_geojson()["features"]) == 58
try:
    import pandas  # noqa: F401
except ModuleNotFoundError:
    pass
else:
    assert len(plotly_data.iris()) == 150
        """,
    ),
    (
        'pluggy-smoke',
        r"""
import pluggy

hookspec = pluggy.HookspecMarker("staticpython")
hookimpl = pluggy.HookimplMarker("staticpython")

class Spec:
    @hookspec
    def answer(self, value):
        pass

class Plugin:
    @hookimpl
    def answer(self, value):
        return value + 1

manager = pluggy.PluginManager("staticpython")
manager.add_hookspecs(Spec)
manager.register(Plugin())
assert manager.hook.answer(value=41) == [42]
        """,
    ),
    (
        'portalocker-smoke',
        r"""
import os
import tempfile
import portalocker

fd, path = tempfile.mkstemp()
os.close(fd)
try:
    with portalocker.Lock(path, "w", timeout=1) as locked:
        locked.write("locked")
    with open(path, encoding="utf-8") as handle:
        assert handle.read() == "locked"
finally:
    os.unlink(path)
        """,
    ),
    (
        'prometheus-client-smoke',
        r"""
from prometheus_client import CollectorRegistry, Counter, Gauge, generate_latest

registry = CollectorRegistry()
counter = Counter("staticpython_requests_total", "StaticPython requests", registry=registry)
gauge = Gauge("staticpython_queue_depth", "StaticPython queue depth", registry=registry)

counter.inc()
counter.inc(2)
gauge.set(7)

payload = generate_latest(registry).decode("utf-8")
assert "staticpython_requests_total 3.0" in payload
assert "staticpython_queue_depth 7.0" in payload
        """,
    ),
    (
        'prompt-toolkit-smoke',
        r"""
from prompt_toolkit.document import Document
from prompt_toolkit.formatted_text import HTML, to_formatted_text
from prompt_toolkit.validation import ValidationError, Validator

text = Document("hello world", cursor_position=5)
assert text.current_line_before_cursor == "hello"
assert to_formatted_text(HTML("<b>demo</b>"))[0][1] == "demo"
class NonEmpty(Validator):
    def validate(self, document):
        if not document.text:
            raise ValidationError(message="empty")
NonEmpty().validate(Document("x"))
        """,
    ),
    (
        'propcache-smoke',
        r"""
from propcache import cached_property, under_cached_property

class Demo:
    def __init__(self):
        self.calls = 0
        self._cache = {}

    @cached_property
    def value(self):
        self.calls += 1
        return self.calls

    @under_cached_property
    def under_value(self):
        self.calls += 10
        return self.calls

obj = Demo()
assert obj.value == 1
assert obj.value == 1
assert obj.under_value == 11
assert obj.under_value == 11
        """,
    ),
    (
        'protobuf-smoke',
        r"""
import importlib.util

from google._upb import _message
from google.protobuf.internal import api_implementation
from google.protobuf.struct_pb2 import Struct

assert importlib.util.find_spec("google._upb._message").origin == "built-in"
assert _message.__name__ == "google._upb._message"
assert api_implementation.Type() == "upb", api_implementation.Type()

payload = Struct()
payload["answer"] = 42
payload["name"] = "codex"
clone = Struct()
clone.ParseFromString(payload.SerializeToString())
assert clone["answer"] == 42 and clone["name"] == "codex", clone
        """,
    ),
    (
        'psutil-smoke',
        r"""
import os
import importlib.util
import psutil

assert importlib.util.find_spec("psutil._psutil_windows").origin == "built-in"
assert psutil.cpu_count() is None or psutil.cpu_count() >= 1
assert psutil.virtual_memory().total > 0
proc = psutil.Process(os.getpid())
assert proc.pid == os.getpid()
assert proc.name()
assert proc.memory_info().rss >= 0
assert isinstance(psutil.net_if_addrs(), dict)
        """,
    ),
    (
        'pure-eval-smoke',
        r"""
import ast
from pure_eval import CannotEval, Evaluator, group_expressions, is_expression_interesting

items = [1, 2, 3]
evaluator = Evaluator({"value": 10, "items": items, "len": len})
expr = ast.parse("value + len(items) + items[0]", mode="eval").body
assert evaluator[expr] == 14

try:
    evaluator[ast.parse("setattr(items, 'x', 1)", mode="eval").body]
except CannotEval:
    pass
else:
    raise AssertionError("unsafe call unexpectedly evaluated")

first = ast.parse("items[0]", mode="eval").body
second = ast.parse("items[0]", mode="eval").body
groups = group_expressions([(first, 1), (second, 1)])
assert len(groups) == 1
assert len(groups[0][0]) == 2
assert not is_expression_interesting(ast.parse("123", mode="eval").body, 123)
        """,
    ),
    (
        'py-smoke',
        r"""
import py

path = py.path.local(".")
assert path.basename
tmp = py.path.local.mkdtemp()
try:
    child = tmp.join("demo.txt")
    child.write("staticpython", ensure=True)
    assert child.read() == "staticpython"
    assert child.check(file=1)
finally:
    tmp.remove(rec=1)
        """,
    ),
    (
        'pyasn1-smoke',
        r"""
from pyasn1.codec.der.decoder import decode
from pyasn1.codec.der.encoder import encode
from pyasn1.type.univ import Integer

encoded = encode(Integer(42))
decoded, rest = decode(encoded, asn1Spec=Integer())
assert int(decoded) == 42
assert rest == b""
        """,
    ),
    (
        'pyasn1-modules-smoke',
        r"""
from pyasn1_modules import rfc5280

assert rfc5280.Certificate.componentType
assert rfc5280.id_ce_basicConstraints.prettyPrint() == "2.5.29.19"
assert rfc5280.Version("v3") == 2
        """,
    ),
    (
        'pycparser-smoke',
        r"""
from pycparser import c_ast, c_generator, c_parser

parser = c_parser.CParser()
tree = parser.parse("int add(int a, int b) { return a + b; }")
functions = [node for _, node in tree.children() if isinstance(node, c_ast.FuncDef)]
assert len(functions) == 1
assert functions[0].decl.name == "add"
generated = c_generator.CGenerator().visit(tree)
assert "int add" in generated
assert "return a + b;" in generated
        """,
    ),
    (
        'pygments-smoke',
        r"""
from pygments import highlight
from pygments.formatters import HtmlFormatter
from pygments.lexers import PythonLexer, get_lexer_by_name

html = highlight("print('x')\n", PythonLexer(), HtmlFormatter())
assert "highlight" in html and "print" in html
assert get_lexer_by_name("python").name == "Python"
        """,
    ),
    (
        'pymongo-smoke',
        r"""
from bson import BSON, ObjectId
from pymongo import MongoClient
from pymongo.uri_parser import parse_uri

oid = ObjectId()
assert ObjectId(str(oid)) == oid
encoded = BSON.encode({"name": "staticpython", "value": 13})
assert BSON(encoded).decode()["value"] == 13
parsed = parse_uri("mongodb://localhost:27017/test")
assert parsed["database"] == "test"
client = MongoClient("mongodb://localhost:27017", connect=False)
assert client.get_database("test").name == "test"
client.close()
        """,
    ),
    (
        'pymysql-smoke',
        r"""
import pymysql
from pymysql.converters import escape_string
from pymysql.cursors import DictCursor

assert pymysql.VERSION_STRING
assert escape_string("a'b") == "a\\'b"
assert DictCursor.__name__ == "DictCursor"
        """,
    ),
    (
        'pyparsing-smoke',
        r"""
from pyparsing import ParseException, Suppress, Word, alphas, delimited_list, nums

integer = Word(nums).set_parse_action(lambda tokens: int(tokens[0]))
record = Word(alphas)("name") + Suppress(":") + delimited_list(integer)("values")
result = record.parse_string("codex:1,2,3", parse_all=True)
assert result.as_dict() == {"name": "codex", "values": [1, 2, 3]}
try:
    record.parse_string("codex:not-a-number", parse_all=True)
except ParseException:
    pass
else:
    raise AssertionError("pyparsing accepted an invalid record")
        """,
    ),
    (
        'pypdf-smoke',
        r"""
import io
from pypdf import PdfReader, PdfWriter

writer = PdfWriter()
writer.add_blank_page(width=72, height=144)
buffer = io.BytesIO()
writer.write(buffer)
buffer.seek(0)
reader = PdfReader(buffer)
assert len(reader.pages) == 1
assert float(reader.pages[0].mediabox.height) == 144.0
        """,
    ),
    (
        'pyperclip-smoke',
        r"""
import pyperclip

assert isinstance(pyperclip.is_available(), bool)
try:
    pyperclip.determine_clipboard()
except Exception as exc:
    assert exc.__class__.__name__ in {"PyperclipException", "RuntimeError"}
        """,
    ),
    (
        'pyrsistent-smoke',
        r"""
from pyrsistent import freeze, m, pmap, pvector, thaw

vec = pvector([1, 2, 3])
assert vec.append(4).tolist() == [1, 2, 3, 4]
assert vec.tolist() == [1, 2, 3]

mapping = pmap({"name": "staticpython"}).set("answer", 42)
assert mapping["answer"] == 42
assert m(a=1).transform(["a"], lambda value: value + 1)["a"] == 2

frozen = freeze({"items": [1, 2], "meta": {"ok": True}})
assert thaw(frozen) == {"items": [1, 2], "meta": {"ok": True}}
        """,
    ),
    (
        'pytest-smoke',
        r"""
import tempfile
from pathlib import Path

import pytest

with tempfile.TemporaryDirectory() as temp_dir:
    root = Path(temp_dir)
    (root / "conftest.py").write_text(
        "import pytest\n@pytest.fixture()\ndef answer():\n    return 42\n",
        encoding="utf-8",
    )
    (root / "test_sample.py").write_text(
        "import pytest\n\n@pytest.mark.parametrize('value', [1, 2, 3])\ndef test_ok(answer, value):\n    assert answer + value in {43, 44, 45}\n",
        encoding="utf-8",
    )
    result = pytest.main([str(root), "-q", "-p", "no:cacheprovider"])
    assert result == 0
        """,
    ),
    (
        'python-multipart-smoke',
        r"""
from multipart.multipart import parse_options_header
from python_multipart.multipart import MultipartParser

value, options = parse_options_header('form-data; name="file"; filename="demo.txt"')
assert value == b"form-data"
assert options[b"name"] == b"file"
assert options[b"filename"] == b"demo.txt"
assert MultipartParser is not None
        """,
    ),
    (
        'python-json-logger-smoke',
        r"""
import io
import json
import logging

from pythonjsonlogger import jsonlogger

stream = io.StringIO()
handler = logging.StreamHandler(stream)
handler.setFormatter(jsonlogger.JsonFormatter())

logger = logging.getLogger("staticpython.pythonjsonlogger")
logger.handlers[:] = []
logger.setLevel(logging.INFO)
logger.propagate = False
logger.addHandler(handler)

logger.info("hello", extra={"answer": 42})
handler.flush()
payload = json.loads(stream.getvalue())

assert payload["message"] == "hello"
assert payload["answer"] == 42
        """,
    ),
    (
        'pytokens-smoke',
        r"""
from pytokens import TokenType, tokenize

tokens = list(tokenize("value = 42\n"))
assert tokens[0].type is TokenType.identifier
assert tokens[0].start_index == 0
assert any(token.type is TokenType.number for token in tokens)
        """,
    ),
    (
        'pytz-smoke',
        r"""
from datetime import datetime

import pytz

utc = pytz.timezone("UTC")
value = utc.localize(datetime(2026, 1, 1, 0, 0, 0))
assert value.utcoffset().total_seconds() == 0
assert pytz.utc.zone == "UTC"
shanghai = pytz.timezone("Asia/Shanghai")
assert shanghai.zone == "Asia/Shanghai"
eastern = pytz.timezone("US/Eastern")
assert eastern.zone == "US/Eastern"
        """,
    ),
    (
        'pyzmq-smoke',
        r"""
import asyncio
import importlib.util

import zmq
import zmq.asyncio
from zmq.utils import z85

assert importlib.util.find_spec("zmq.backend.cython._zmq").origin == "built-in"
assert zmq.has("curve"), "CURVE support should be enabled"
assert not zmq.has("ipc"), "Windows pyzmq static build should disable IPC with select poller"

payload = b"0123456789abcdef"
encoded = z85.encode(payload)
assert z85.decode(encoded) == payload

context = zmq.Context()
left = context.socket(zmq.PAIR)
right = context.socket(zmq.PAIR)
endpoint = "inproc://staticpython-pyzmq-sync"
left.bind(endpoint)
right.connect(endpoint)
poller = zmq.Poller()
poller.register(right, zmq.POLLIN)
left.send_multipart([b"alpha", b"beta"])
events = dict(poller.poll(1000))
assert events.get(right) == zmq.POLLIN
assert right.recv_multipart() == [b"alpha", b"beta"]
frame = zmq.Frame(b"frame-data")
assert bytes(frame) == b"frame-data"
assert right.getsockopt(zmq.TYPE) == zmq.PAIR
left.close(0)
right.close(0)
context.term()

async def _probe_asyncio():
    actx = zmq.asyncio.Context()
    async_left = actx.socket(zmq.PAIR)
    async_right = actx.socket(zmq.PAIR)
    async_endpoint = "inproc://staticpython-pyzmq-async"
    async_left.bind(async_endpoint)
    async_right.connect(async_endpoint)
    await async_left.send_json({"answer": 42})
    data = await async_right.recv_json()
    assert data == {"answer": 42}
    async_left.close(0)
    async_right.close(0)
    actx.term()

asyncio.run(_probe_asyncio())
        """,
    ),
    (
        'rapidfuzz-smoke',
        r"""
from rapidfuzz import fuzz, process
from rapidfuzz.distance import Levenshtein

assert fuzz.ratio("static python", "static python") == 100
assert fuzz.partial_ratio("single-file python", "python") == 100
assert Levenshtein.distance("kitten", "sitting") == 3
choice = process.extractOne("static", ["dynamic", "static", "frozen"])
assert choice[0] == "static", choice
        """,
    ),
    (
        'redis-smoke',
        r"""
import redis
from redis.connection import parse_url

client = redis.Redis.from_url("redis://localhost:6379/2?decode_responses=True")
kwargs = client.connection_pool.connection_kwargs
assert kwargs["host"] == "localhost"
assert kwargs["db"] == 2
parsed = parse_url("redis://:pass@example.com:6380/1")
assert parsed["host"] == "example.com" and parsed["port"] == 6380 and parsed["db"] == 1
        """,
    ),
    (
        'regex-smoke',
        r"""
import regex
import importlib.util

assert importlib.util.find_spec("regex._regex").origin == "built-in"
match = regex.search(r"(?P<word>\p{Letter}+)", "abc 123")
assert match and match.group("word") == "abc", match
assert regex.findall(r"\X", "a\u0301b") == ["a\u0301", "b"]
assert regex.sub(r"(\w+)", r"[\1]", "codex") == "[codex]"
        """,
    ),
    (
        'requests-smoke',
        r"""
from collections import OrderedDict
import requests
from requests import Request, Session
from requests.cookies import RequestsCookieJar

request = requests.Request("POST", "https://example.com/api", params=OrderedDict([("a", "1"), ("b", "2")]), json={"ok": True})
prepared = request.prepare()
assert prepared.method == "POST"
assert prepared.url == "https://example.com/api?a=1&b=2"
assert prepared.body == b'{"ok": true}'
session = requests.Session()
assert "https://" in session.adapters

cookie_jar = RequestsCookieJar()
cookie_jar.set("token", "abc", domain="example.com", path="/")
session = Session()
session.cookies = cookie_jar
prepared_with_cookie = session.prepare_request(Request("GET", "https://example.com/demo", params={"x": "1"}))
assert prepared_with_cookie.headers["Cookie"] == "token=abc"
assert prepared_with_cookie.url == "https://example.com/demo?x=1"

response = requests.Response()
response.status_code = 200
response._content = b'{"ok": true}'
response.headers["Content-Type"] = "application/json"
assert response.json() == {"ok": True}
        """,
    ),
    (
        'requests-oauthlib-smoke',
        r"""
from requests_oauthlib import OAuth1

auth = OAuth1("client-key", client_secret="client-secret")
assert auth.client.client_key == "client-key"
assert auth.client.client_secret == "client-secret"
        """,
    ),
    (
        'requests-toolbelt-smoke',
        r"""
import requests
from requests_toolbelt import MultipartEncoder
from requests_toolbelt.sessions import BaseUrlSession

encoder = MultipartEncoder(
    fields={
        "field": "value",
        "file": ("demo.txt", b"payload", "text/plain"),
    }
)
body = encoder.to_string()
assert encoder.content_type.startswith("multipart/form-data; boundary=")
assert b'name="field"' in body
assert b'filename="demo.txt"' in body
assert b"payload" in body

session = BaseUrlSession(base_url="https://example.com/api/")
request = requests.Request("GET", "users", params={"q": "codex"})
prepared = session.prepare_request(request)
assert prepared.url == "https://example.com/api/users?q=codex"
        """,
    ),
    (
        'responses-smoke',
        r"""
import requests
import responses


@responses.activate
def run():
    responses.add(
        responses.GET,
        "https://example.com/api",
        json={"ok": True, "items": [1, 2]},
        status=200,
    )
    response = requests.get("https://example.com/api", timeout=1)
    assert response.status_code == 200
    assert response.json() == {"ok": True, "items": [1, 2]}
    assert len(responses.calls) == 1
    assert responses.calls[0].request.url == "https://example.com/api"


run()
        """,
    ),
    (
        'rfc3339-validator-smoke',
        r"""
from rfc3339_validator import validate_rfc3339

assert validate_rfc3339("2024-01-01T00:00:00Z")
assert validate_rfc3339("2024-01-01T00:00:00+08:00")
assert not validate_rfc3339("2024-13-01T00:00:00Z")
assert not validate_rfc3339("not-a-timestamp")
        """,
    ),
    (
        'rfc3986-validator-smoke',
        r"""
from rfc3986_validator import validate_rfc3986

assert validate_rfc3986("https://example.com/staticpython")
assert validate_rfc3986("mailto:test@example.com", rule="URI")
assert not validate_rfc3986("not a uri")
        """,
    ),
    (
        'rich-smoke',
        r"""
import io
from rich.console import Console
from rich.table import Table
from rich.text import Text

buffer = io.StringIO()
console = Console(file=buffer, force_terminal=False, color_system=None, width=80)
table = Table("name", "value")
table.add_row("codex", "42")
console.print(table)
console.print(Text("plain", style="bold"))
output = buffer.getvalue()
assert "codex" in output and "42" in output
assert "plain" in output
        """,
    ),
    (
        'rsa-smoke',
        r"""
import rsa

public_key, private_key = rsa.newkeys(512)
message = b"staticpython"
signature = rsa.sign(message, private_key, "SHA-256")
assert rsa.verify(message, signature, public_key) == "SHA-256"
try:
    rsa.verify(message + b"!", signature, public_key)
except rsa.VerificationError:
    pass
else:
    raise AssertionError("rsa accepted a tampered signature")
encrypted = rsa.encrypt(message, public_key)
assert rsa.decrypt(encrypted, private_key) == message
        """,
    ),
    (
        'selenium-smoke',
        r"""
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

options = Options()
options.add_argument("--headless=new")
capabilities = options.to_capabilities()
assert capabilities["browserName"] == "chrome"
assert "--headless=new" in capabilities["goog:chromeOptions"]["args"]
assert By.CSS_SELECTOR == "css selector"
assert Keys.ENTER == "\ue007"
        """,
    ),
    (
        'send2trash-smoke',
        r"""
import tempfile
from pathlib import Path

from send2trash import send2trash

with tempfile.TemporaryDirectory() as temp_dir:
    file_path = Path(temp_dir) / "staticpython-trash.txt"
    file_path.write_text("trash me", encoding="utf-8")
    send2trash(str(file_path))
    assert not file_path.exists()
        """,
    ),
    (
        'shellingham-smoke',
        r"""
import shellingham

try:
    shell = shellingham.detect_shell()
except shellingham.ShellDetectionFailure:
    shell = None
assert shell is None or (isinstance(shell, tuple) and len(shell) == 2)
assert issubclass(shellingham.ShellDetectionFailure, OSError)
        """,
    ),
    (
        'six-smoke',
        r"""
import six

assert six.text_type("demo") == "demo"
assert list(six.moves.range(3)) == [0, 1, 2]
assert six.ensure_text(b"demo") == "demo"
assert six.iteritems({"a": 1}).__next__() == ("a", 1)
        """,
    ),
    (
        'python-slugify-smoke',
        r"""
from slugify import slugify

assert slugify("Static Python: Café déjà vu!") == "static-python-cafe-deja-vu"
assert slugify("影師嗎", allow_unicode=True) == "影師嗎"
assert slugify("影師嗎") == "ying-shi-ma"
        """,
    ),
    (
        'sniffio-smoke',
        r"""
import anyio
import sniffio

async def probe():
    return sniffio.current_async_library()

assert anyio.run(probe) == "asyncio"
try:
    sniffio.current_async_library()
except sniffio.AsyncLibraryNotFoundError:
    pass
else:
    raise AssertionError("sniffio detected an async library outside async context")
        """,
    ),
    (
        'pysocks-smoke',
        r"""
import socks

sock = socks.socksocket()
sock.set_proxy(socks.SOCKS5, "localhost", 1080, username="user", password="pass")
assert sock.proxy[0] == socks.SOCKS5
assert sock.proxy[1] == "localhost"
assert sock.proxy[2] == 1080
assert socks.PROXY_TYPE_HTTP == socks.HTTP
sock.close()
        """,
    ),
    (
        'sortedcontainers-smoke',
        r"""
from sortedcontainers import SortedDict, SortedList

values = SortedList([3, 1, 2])
assert list(values) == [1, 2, 3]
mapping = SortedDict({"b": 2, "a": 1})
assert list(mapping.items()) == [("a", 1), ("b", 2)]
        """,
    ),
    (
        'soupsieve-smoke',
        r"""
import soupsieve
from bs4 import BeautifulSoup

soup = BeautifulSoup("<div><p class='a'>one</p><p>two</p></div>", "html.parser")
selector = soupsieve.compile("p.a")
assert selector.select_one(soup).text == "one"
assert [node.text for node in soupsieve.select("p", soup)] == ["one", "two"]
        """,
    ),
    (
        'sqlalchemy-smoke',
        r"""
import sqlalchemy as sa

engine = sa.create_engine("sqlite:///:memory:")
metadata = sa.MetaData()
table = sa.Table("demo", metadata, sa.Column("id", sa.Integer, primary_key=True), sa.Column("name", sa.String))
metadata.create_all(engine)
with engine.begin() as conn:
    conn.execute(table.insert(), [{"name": "a"}, {"name": "b"}])
    rows = conn.execute(sa.select(table.c.name).order_by(table.c.id)).scalars().all()
    count = conn.scalar(sa.select(sa.func.count()).select_from(table))
assert rows == ["a", "b"]
assert count == 2
assert "demo" in sa.inspect(engine).get_table_names()
        """,
    ),
    (
        'sqlparse-smoke',
        r"""
import sqlparse

statements = sqlparse.split("select 1; select 2;")
assert statements == ["select 1;", "select 2;"]
formatted = sqlparse.format("select * from demo where id=1", keyword_case="upper", reindent=True)
assert "SELECT" in formatted and "FROM" in formatted
        """,
    ),
    (
        'stack-data-smoke',
        r"""
import linecache
import sys
import textwrap
from stack_data import FrameInfo, Line, Options

filename = "<stack_data_smoke>"
source_text = textwrap.dedent(
    '''
    import sys
    def probe():
        alpha = 20
        beta = [1, 2, 3]
        frame = sys._getframe()
        return frame
    '''
)
linecache.cache[filename] = (len(source_text), None, source_text.splitlines(True), filename)
namespace = {"sys": sys}
exec(compile(source_text, filename, "exec"), namespace)
frame = namespace["probe"]()
info = FrameInfo(frame, Options(before=1, after=1))
lines = list(info.lines)
variables = {variable.name: variable.value for variable in info.variables}
assert info.source.filename == filename
assert any(isinstance(line, Line) and line.is_current for line in lines)
assert variables["alpha"] == 20
assert variables["beta"] == [1, 2, 3]
        """,
    ),
    (
        'starlette-smoke',
        r"""
import anyio
from starlette.applications import Starlette
from starlette.responses import PlainTextResponse
from starlette.routing import Route
from starlette.testclient import TestClient


async def homepage(request):
    return PlainTextResponse("staticpython")


app = Starlette(routes=[Route("/", homepage)])
with TestClient(app) as client:
    response = client.get("/")
    assert response.status_code == 200
    assert response.text == "staticpython"

async def run_task():
    return "ok"

assert anyio.run(run_task) == "ok"
        """,
    ),
    (
        'sympy-smoke',
        r"""
import sympy as sp

x = sp.Symbol("x")
assert sp.factor(x**2 - 1) == (x - 1) * (x + 1)
assert sp.integrate(2 * x, x) == x**2
assert sp.solve(sp.Eq(x + 2, 5), x) == [3]
        """,
    ),
    (
        'tabulate-smoke',
        r"""
from tabulate import tabulate

table = tabulate([["codex", 42]], headers=["name", "value"], tablefmt="github")
assert "codex" in table
assert "| name" in table
plain = tabulate([[1, 2], [3, 4]], tablefmt="plain")
assert "1" in plain and "4" in plain
        """,
    ),
    (
        'tenacity-smoke',
        r"""
from tenacity import retry, retry_if_exception_type, stop_after_attempt

calls = []
@retry(stop=stop_after_attempt(3), retry=retry_if_exception_type(ValueError), reraise=True)
def flaky():
    calls.append(1)
    if len(calls) < 2:
        raise ValueError("try again")
    return "ok"

assert flaky() == "ok"
assert len(calls) == 2
        """,
    ),
    (
        'terminado-smoke',
        r"""
from terminado.management import _update_removing
from terminado.websocket import TermSocket

data = {"keep": 1, "drop": 2}
_update_removing(data, {"drop": None, "add": 3})

assert data == {"keep": 1, "add": 3}
assert callable(TermSocket.send_json_message)
assert callable(TermSocket.origin_check)
        """,
    ),
    (
        'text-unidecode-smoke',
        r"""
from text_unidecode import unidecode

assert unidecode("Café déjà vu").startswith("Cafe deja vu")
assert "Ying" in unidecode("影")
        """,
    ),
    (
        'threadpoolctl-smoke',
        r"""
from threadpoolctl import ThreadpoolController, threadpool_info, threadpool_limits

assert isinstance(threadpool_info(), list)
controller = ThreadpoolController()
assert isinstance(controller.info(), list)
with threadpool_limits(limits=1):
    assert isinstance(threadpool_info(), list)
        """,
    ),
    (
        'tinycss2-smoke',
        r"""
from tinycss2 import parse_stylesheet, serialize

rules = parse_stylesheet("h1 { color: red; margin: 0 }", skip_whitespace=True)
assert rules and rules[0].type == "qualified-rule"
assert "color" in serialize(rules[0].content)
        """,
    ),
    (
        'tomlkit-smoke',
        r"""
import tomlkit

doc = tomlkit.parse("[tool.demo]\nanswer = 42\n")
assert doc["tool"]["demo"]["answer"] == 42
doc["tool"]["demo"]["name"] = "codex"
rendered = tomlkit.dumps(doc)
assert 'name = "codex"' in rendered
        """,
    ),
    (
        'tornado-smoke',
        r"""
from tornado.escape import json_decode, json_encode
from tornado.httputil import url_concat

assert json_decode(json_encode({"ok": True})) == {"ok": True}
assert url_concat("https://example.com/api", {"q": "static python"}) == "https://example.com/api?q=static+python"
        """,
    ),
    (
        'tqdm-smoke',
        r"""
from tqdm import tqdm

assert "100%" in tqdm.format_meter(10, 10, 1.0)
bar = tqdm(iterable=[1, 2, 3], disable=True)
assert list(bar) == [1, 2, 3]
        """,
    ),
    (
        'traitlets-smoke',
        r"""
from traitlets import HasTraits, Int, TraitError

class Counter(HasTraits):
    value = Int(0)

seen = []
counter = Counter()
counter.observe(lambda change: seen.append(change["new"]), names="value")
counter.value = 5
assert seen == [5]
try:
    counter.value = "bad"
except TraitError:
    pass
else:
    raise AssertionError("traitlets accepted a non-integer value")
        """,
    ),
    (
        'typer-smoke',
        r"""
import typer
from typer.testing import CliRunner

app = typer.Typer()

@app.command()
def hello(name: str):
    typer.echo(f"hello {name}")

result = CliRunner().invoke(app, ["codex"])
assert result.exit_code == 0
assert "hello codex" in result.output
        """,
    ),
    (
        'typing-extensions-smoke',
        r"""
from typing_extensions import Annotated, Literal, TypedDict, get_args, get_origin

class Demo(TypedDict):
    value: int

assert Demo(value=1)["value"] == 1
annotated = Annotated[int, "meta"]
assert get_origin(annotated) is Annotated
assert get_args(annotated) == (int, "meta")
assert get_args(Literal["a", "b"]) == ("a", "b")
        """,
    ),
    (
        'typing-inspect-smoke',
        r"""
from typing import List, Optional, Union

from typing_inspect import get_args, get_origin, is_optional_type, is_union_type

assert get_origin(List[int]) is list
assert get_args(List[int]) == (int,)
assert is_optional_type(Optional[int])
assert is_union_type(Union[int, str])
        """,
    ),
    (
        'typing-inspection-smoke',
        r"""
from typing import Literal, Union

from typing_inspection.introspection import get_literal_values
from typing_inspection.typing_objects import is_union

assert list(get_literal_values(Literal["a", "b"])) == ["a", "b"]
assert is_union(Union)
        """,
    ),
    (
        'tzdata-smoke',
        r"""
import zoneinfo
import tzdata

zones = tzdata.available_timezones()
assert "UTC" in zones
utc = zoneinfo.ZoneInfo("UTC")
assert utc.key == "UTC"
assert tzdata.open_zoneinfo("UTC").read(4) == b"TZif"
        """,
    ),
    (
        'ujson-smoke',
        r"""
import math
import importlib.util
import ujson

assert importlib.util.find_spec("ujson").origin == "built-in"
payload = {"name": "codex", "items": [1, 2, 3], "value": 1.25}
encoded = ujson.dumps(payload)
decoded = ujson.loads(encoded)
assert decoded == payload, decoded
assert math.isnan(ujson.loads("NaN"))
try:
    ujson.loads("{broken")
except ValueError:
    pass
else:
    raise AssertionError("ujson did not reject malformed JSON")
        """,
    ),
    (
        'urllib3-smoke',
        r"""
import urllib3
from urllib3.util import Retry, parse_url

url = parse_url("https://example.com:443/path?q=1")
assert url.scheme == "https" and url.host == "example.com" and url.port == 443
retry = Retry(total=3, status_forcelist=[500])
assert retry.total == 3 and retry.is_retry("GET", 500)
manager = urllib3.PoolManager(num_pools=1, maxsize=1)
assert manager.connection_pool_kw["maxsize"] == 1
        """,
    ),
    (
        'uvicorn-smoke',
        r"""
from uvicorn.config import Config
from uvicorn.importer import import_from_string

config = Config("uvicorn.main:main", loop="asyncio", http="h11", lifespan="off", log_config=None)
assert config.app == "uvicorn.main:main"
assert import_from_string("uvicorn.config:Config") is Config
        """,
    ),
    (
        'wcwidth-smoke',
        r"""
from wcwidth import wcwidth, wcswidth

assert wcwidth("a") == 1
assert wcwidth("́") == 0
assert wcswidth("abc") == 3
assert wcswidth("中文") == 4
        """,
    ),
    (
        'webencodings-smoke',
        r"""
from webencodings import ascii_lower, lookup

assert ascii_lower("UTF-8") == "utf-8"
assert lookup("utf-8").name == "utf-8"
        """,
    ),
    (
        'websocket-client-smoke',
        r"""
from websocket import ABNF
from websocket._url import parse_url

frame = ABNF.create_frame("hello", ABNF.OPCODE_TEXT)
frame.validate(skip_utf8_validation=False)
assert frame.opcode == ABNF.OPCODE_TEXT
assert frame.data == b"hello"
host, port, resource, secure = parse_url("wss://example.com/chat")
assert (host, port, resource, secure) == ("example.com", 443, "/chat", True)
        """,
    ),
    (
        'websockets-smoke',
        r"""
import importlib.util

import websockets
import websockets.frames
import websockets.speedups as speedups
from websockets.uri import parse_uri

assert importlib.util.find_spec("websockets.speedups").origin == "built-in"
assert speedups.apply_mask(b"abcd", b"\x01\x02\x03\x04") == bytes([0x60, 0x60, 0x60, 0x60])
assert websockets.frames.apply_mask is speedups.apply_mask
uri = parse_uri("wss://example.com/chat")
assert uri.secure is True and uri.host == "example.com" and uri.path == "/chat", uri
        """,
    ),
    (
        'werkzeug-smoke',
        r"""
from werkzeug.datastructures import MultiDict
from werkzeug.routing import Map, Rule
from werkzeug.wrappers import Request, Response

mapping = Map([Rule("/hello/<name>", endpoint="hello")])
adapter = mapping.bind("example.com")
assert adapter.match("/hello/codex") == ("hello", {"name": "codex"})
assert MultiDict([("a", "1"), ("a", "2")]).getlist("a") == ["1", "2"]
response = Response("ok", status=201)
assert response.status_code == 201 and response.get_data(as_text=True) == "ok"
        """,
    ),
    (
        'win32-setctime-smoke',
        r"""
import os
import tempfile
import win32_setctime

assert isinstance(win32_setctime.SUPPORTED, bool)
fd, path = tempfile.mkstemp()
os.close(fd)
try:
    if win32_setctime.SUPPORTED:
        win32_setctime.setctime(path, 1_700_000_000)
    assert os.path.exists(path)
finally:
    os.unlink(path)
        """,
    ),
    (
        'wsproto-smoke',
        r"""
from wsproto.frame_protocol import CloseReason, Opcode
from wsproto.utilities import generate_accept_token

token = generate_accept_token(b"dGhlIHNhbXBsZSBub25jZQ==")
assert token == b"s3pPLMBiTxaQ9kYGzzhZRbK+xOo="
assert CloseReason.NORMAL_CLOSURE == 1000
assert Opcode.TEXT == 0x1
        """,
    ),
    (
        'xlsxwriter-smoke',
        r"""
import io
import zipfile
import xlsxwriter

buffer = io.BytesIO()
workbook = xlsxwriter.Workbook(buffer, {"in_memory": True})
sheet = workbook.add_worksheet("Data")
sheet.write(0, 0, "codex")
sheet.write_formula(0, 1, "=SUM(1,2)")
workbook.close()
buffer.seek(0)
with zipfile.ZipFile(buffer) as archive:
    names = archive.namelist()
assert "xl/workbook.xml" in names
assert any(name.startswith("xl/worksheets/") for name in names)
        """,
    ),
    (
        'xmltodict-smoke',
        r"""
import xmltodict

data = xmltodict.parse("<root><item id='1'>ok</item></root>")
assert data["root"]["item"]["@id"] == "1"
assert data["root"]["item"]["#text"] == "ok"
xml = xmltodict.unparse(data)
assert "<root>" in xml and "<item id=\"1\">ok</item>" in xml

namespaced = xmltodict.parse(
    '<root xmlns:a="urn:test"><item id="1">ok</item><item id="2">fine</item><a:meta>yes</a:meta></root>',
    process_namespaces=True,
    namespaces={"urn:test": "ns"},
    force_list=("item",),
)
assert len(namespaced["root"]["item"]) == 2
assert namespaced["root"]["item"][1]["@id"] == "2"
assert namespaced["root"]["ns:meta"] == "yes"

fragment = xmltodict.unparse(
    {"root": {"item": [{"@id": "1", "#text": "ok"}]}},
    full_document=False,
)
assert '<item id="1">ok</item>' in fragment
        """,
    ),
    (
        'yaml-smoke',
        r"""
import yaml

data = yaml.safe_load("items:\n  - 1\n  - 2\nname: codex\n")
assert data == {"items": [1, 2], "name": "codex"}
rendered = yaml.safe_dump(data, sort_keys=True)
assert "items:" in rendered and "name: codex" in rendered
        """,
    ),
    (
        'bokeh-smoke',
        r"""
from bokeh.embed import json_item
from bokeh.plotting import figure

plot = figure(title="StaticPython", width=320, height=240)
plot.line([1, 2, 3], [2, 4, 6])
item = json_item(plot, "target")
assert item["target_id"] == "target"
assert "doc" in item
        """,
    ),
    (
        'dash-smoke',
        r"""
from dash import Dash, dcc, html

app = Dash("staticpython_dash_smoke", server=False)
app.layout = html.Div([html.H1("StaticPython"), dcc.Graph(id="plot")])
assert app.layout.children[0].children == "StaticPython"
assert app.layout.children[1].id == "plot"
        """,
    ),
    (
        'dialite-smoke',
        r"""
import dialite

with dialite.NoDialogs():
    supported = dialite.is_supported()
assert supported in (True, False)
assert hasattr(dialite, "inform")
        """,
    ),
    (
        'flexx-smoke',
        r"""
import flexx
from flexx import event, flx

data = event.Dict(answer=42)
assert data.answer == 42
assert flexx.__version__
assert hasattr(flx, "Widget")
        """,
    ),
    (
        'opengl-smoke',
        r"""
from OpenGL import GL

assert GL.GL_TRIANGLES == 0x0004
assert callable(GL.glGetError)
        """,
    ),
    (
        'imgui-smoke',
        r"""
import imgui
import importlib

def assert_builtin_native(name):
    module = importlib.import_module(name)
    spec = getattr(module, "__spec__", None)
    origin = getattr(spec, "origin", None)
    file_name = getattr(module, "__file__", None)
    assert origin == "built-in", (name, origin, file_name)

ctx = imgui.create_context()
try:
    assert_builtin_native("imgui.core")
    assert imgui.__version__
    assert imgui.VERTEX_SIZE > 0
    io = imgui.get_io()
    assert io is not None
    try:
        assert_builtin_native("imgui.internal")
    except ModuleNotFoundError:
        pass
finally:
    imgui.destroy_context(ctx)
        """,
    ),
    (
        'glfw-smoke',
        r"""
import glfw
import importlib

def assert_builtin_native(name):
    module = importlib.import_module(name)
    spec = getattr(module, "__spec__", None)
    origin = getattr(spec, "origin", None)
    file_name = getattr(module, "__file__", None)
    assert origin == "built-in", (name, origin, file_name)

assert_builtin_native("glfw._glfw")
assert glfw.__version__
assert glfw.get_version()[:2] >= (3, 4)
assert glfw.TRUE == 1
assert glfw.FALSE == 0
assert glfw.KEY_A > 0
assert glfw.VISIBLE > 0
assert callable(glfw.init)
assert callable(glfw.create_window)
        """,
    ),
    (
        'pscript-smoke',
        r"""
from pscript import py2js

js = py2js("def add(a, b):\n    return a + b\n")
assert "add" in js and "return" in js and "a + b" in js
        """,
    ),
    (
        'pyglet-smoke',
        r"""
import pyglet
from pyglet.event import EventDispatcher
from pyglet.math import Vec2

vector = Vec2(3, 4)
length = vector.length() if callable(vector.length) else vector.length
assert length == 5

class Demo(EventDispatcher):
    pass

Demo.register_event_type("on_ping")
demo = Demo()
events = []

@demo.event
def on_ping(value):
    events.append(value)

demo.dispatch_event("on_ping", 7)
assert events == [7]
assert pyglet.version
        """,
    ),
    (
        'pyfltk-smoke',
        r"""
import fltk
import importlib

def assert_builtin_native(name):
    module = importlib.import_module(name)
    spec = getattr(module, "__spec__", None)
    origin = getattr(spec, "origin", None)
    file_name = getattr(module, "__file__", None)
    assert origin == "built-in", (name, origin, file_name)

assert_builtin_native("fltk._fltk")
assert fltk.__version__
assert hasattr(fltk, "Fl_Window")
assert hasattr(fltk, "Fl_Button")
assert fltk.Window is fltk.Fl_Window
assert fltk.Input is fltk.Fl_Input
assert fltk.Button is fltk.Fl_Button
assert callable(fltk.run)
        """,
    ),
    (
        'dearpygui-smoke',
        r"""
import dearpygui
import dearpygui.dearpygui as dpg
import importlib

def assert_builtin_native(name):
    module = importlib.import_module(name)
    spec = getattr(module, "__spec__", None)
    origin = getattr(spec, "origin", None)
    file_name = getattr(module, "__file__", None)
    assert origin == "built-in", (name, origin, file_name)

assert_builtin_native("dearpygui._dearpygui")
assert dearpygui.__version__
assert hasattr(dpg, "create_context")
assert hasattr(dpg, "create_viewport")
dpg.create_context()
dpg.destroy_context()
        """,
    ),
    (
        'pywebio-smoke',
        r"""
import pywebio
from pywebio.output import put_text
from pywebio.session import local

assert pywebio.__version__
assert callable(put_text)
local.answer = 42
assert local.answer == 42
        """,
    ),
    (
        'pystray-smoke',
        r"""
from PIL import Image
from pystray import Icon, Menu, MenuItem

image = Image.new("RGBA", (16, 16), (255, 0, 0, 255))
item = MenuItem("Quit", lambda icon, item: None)
menu = Menu(item)
icon = Icon("staticpython", image, "StaticPython", menu)
assert icon.name == "staticpython"
assert icon.title == "StaticPython"
assert tuple(icon.icon.size) == (16, 16)
assert len(tuple(menu)) == 1
        """,
    ),
    (
        'remi-smoke',
        r"""
from remi import gui

container = gui.VBox(width=120, height=80)
button = gui.Button("Run")
button.style["color"] = "red"
container.append(button, "button")
assert container.children["button"] is button
assert button.get_text() == "Run"
        """,
    ),
    (
        'retrying-smoke',
        r"""
from retrying import Retrying

attempts = []

def work():
    attempts.append("called")
    return "ok"

assert Retrying(stop_max_attempt_number=1).call(work) == "ok"
assert attempts == ["called"]
        """,
    ),
    (
        'user-agents-smoke',
        r"""
from ua_parser import user_agent_parser
from user_agents import parse

text = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
ua = parse(text)
assert ua.browser.family
parsed = user_agent_parser.Parse(text)
assert parsed["user_agent"]["family"]
        """,
    ),
    (
        'webruntime-smoke',
        r"""
import webruntime

assert hasattr(webruntime, "launch")
assert hasattr(webruntime, "BaseRuntime")
assert "browser" in webruntime._runtimes
        """,
    ),
    (
        'xyzservices-smoke',
        r"""
import xyzservices.providers as xyz

assert xyz.OpenStreetMap.Mapnik.name
        """,
    ),
    (
        'yarl-smoke',
        r"""
from yarl import URL

url = URL("https://example.com") / "api" % {"q": "static python"}
assert str(url) == "https://example.com/api?q=static+python"
assert url.with_scheme("http").scheme == "http"
assert URL("/a/b").parts == ("/", "a", "b")
        """,
    ),
    (
        'zipp-smoke',
        r"""
import io
import zipfile

from zipp import CompleteDirs, Path

assert list(CompleteDirs._implied_dirs(["demo/pkg/module.py"])) == ["demo/pkg/", "demo/"]
buffer = io.BytesIO()
with zipfile.ZipFile(buffer, "w") as archive:
    archive.writestr("pkg/data.txt", "ok")
with zipfile.ZipFile(buffer) as archive:
    root = Path(archive)
    assert (root / "pkg" / "data.txt").read_text() == "ok"
        """,
    ),
]


SUBPROCESS_TESTS = [
    {
        "kind": "module",
        "name": "crypto-selftest",
        "module": "Crypto.SelfTest",
        "timeout": 900,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_CRYPTO",
    },
    {
        "kind": "script",
        "name": "jupyter-server-runtime",
        "script": "scripts/jupyter_runtime.py",
        "args": ["--target", "server", "--timeout", "240"],
        "timeout": 300,
    },
    {
        "kind": "script",
        "name": "jupyterlab-runtime",
        "script": "scripts/jupyter_runtime.py",
        "args": ["--target", "lab", "--timeout", "240"],
        "timeout": 300,
    },
    {
        "kind": "script",
        "name": "libui-smoke",
        "script": "assets/overlay/libui_smoke_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "libui-unittest",
        "script": "assets/overlay/Lib/test/test_libui.py",
        "timeout": 600,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "libui-gui-unittest",
        "script": "assets/overlay/Lib/test/test_libui_gui.py",
        "timeout": 600,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "imgui-runtime",
        "script": "assets/overlay/imgui_runtime_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "imgui-pyglet-runtime",
        "script": "assets/overlay/imgui_pyglet_runtime_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "imgui-glfw-runtime",
        "script": "assets/overlay/imgui_glfw_runtime_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "pyfltk-runtime",
        "script": "assets/overlay/pyfltk_runtime_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "dearpygui-runtime",
        "script": "assets/overlay/dearpygui_runtime_test.py",
        "timeout": 180,
        "skip_env": "STATICPYTHON_VERIFY_SKIP_GUI",
    },
    {
        "kind": "script",
        "name": "nbconvert-runtime",
        "script": "scripts/nbconvert_runtime.py",
        "timeout": 600,
    },
    {
        "kind": "script",
        "name": "notebook-runtime",
        "script": "scripts/jupyter_runtime.py",
        "args": ["--target", "notebook", "--timeout", "240"],
        "timeout": 300,
    },
    {
        "kind": "script",
        "name": "nest-asyncio-runtime",
        "script": "scripts/nest_asyncio_runtime.py",
        "timeout": 120,
    },
]


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _subprocess_command(step: dict) -> list[str]:
    if step["kind"] == "module":
        return [sys.executable, "-m", step["module"], *[str(arg) for arg in step.get("args", [])]]
    if step["kind"] == "script":
        return [sys.executable, str(_repo_root() / step["script"]), *[str(arg) for arg in step.get("args", [])]]
    raise RuntimeError(f"unsupported subprocess test kind: {step['kind']!r}")


def run_subprocess_test(step: dict) -> tuple[str, BaseException | None, str]:
    name = step["name"]
    skip_env = step.get("skip_env")
    if skip_env and os.environ.get(skip_env):
        print(f"[staticpython-full-verify] {name}: skipped because {skip_env}=1", flush=True)
        return name, None, ""

    command = _subprocess_command(step)
    print(f"[staticpython-full-verify] {name}: running {' '.join(command)}", flush=True)
    try:
        with tempfile.TemporaryDirectory(prefix=f"staticpython-verify-{name}-") as cwd:
            completed = subprocess.run(
                command,
                cwd=cwd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=float(step.get("timeout", 240)),
            )
    except BaseException as exc:
        return name, exc, traceback.format_exc()

    if completed.stdout:
        print(completed.stdout, end="" if completed.stdout.endswith("\n") else "\n", flush=True)
    if completed.stderr:
        print(completed.stderr, end="" if completed.stderr.endswith("\n") else "\n", flush=True)
    if completed.returncode != 0:
        return name, RuntimeError(f"command exited with code {completed.returncode}"), (
            f"command: {command!r}\nstdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    print(f"[staticpython-full-verify] {name}: passed", flush=True)
    return name, None, ""


def main() -> int:
    failures = []
    for name, code in SMOKE_TESTS:
        print(f"[staticpython-full-verify] {name}: running", flush=True)
        module_name = "__staticpython_full_verify__"
        module = types.ModuleType(module_name)
        module.__file__ = f"<staticpython-full-verify:{name}>"
        previous_module = sys.modules.get(module_name)
        sys.modules[module_name] = module
        try:
            exec(compile(code, f"<staticpython-full-verify:{name}>", "exec"), module.__dict__)
        except BaseException as exc:
            failures.append((name, exc, traceback.format_exc()))
            print(f"[staticpython-full-verify] {name}: failed: {exc!r}", flush=True)
        else:
            print(f"[staticpython-full-verify] {name}: passed", flush=True)
        finally:
            if previous_module is None:
                sys.modules.pop(module_name, None)
            else:
                sys.modules[module_name] = previous_module
    for step in SUBPROCESS_TESTS:
        name, exc, details = run_subprocess_test(step)
        if exc is not None:
            failures.append((name, exc, details))
            print(f"[staticpython-full-verify] {name}: failed: {exc!r}", flush=True)
    if failures:
        print(f"[staticpython-full-verify] {len(failures)} failure(s)", flush=True)
        for name, exc, tb in failures:
            print(f"[staticpython-full-verify] FAILURE {name}: {exc!r}", flush=True)
            print(tb, flush=True)
        return 1
    print(
        f"[staticpython-full-verify] all {len(SMOKE_TESTS)} inline and {len(SUBPROCESS_TESTS)} subprocess smoke test(s) passed",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
